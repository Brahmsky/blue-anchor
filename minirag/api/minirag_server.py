# pyright: reportMissingImports=false, reportGeneralTypeIssues=false, reportAttributeAccessIssue=false, reportAssignmentType=false, reportArgumentType=false, reportReturnType=false, reportCallIssue=false, reportOptionalMemberAccess=false, reportPossiblyUnboundVariable=false, reportOperatorIssue=false, reportOptionalOperand=false
import asyncio
import contextlib
import csv

from fastapi import FastAPI, HTTPException, File, UploadFile, Form, Request, Query
from fastapi.responses import Response

# Backend (Python)
# Add this to store progress globally
from typing import Dict, Optional, Any
from datetime import datetime
import argparse
import importlib
import logging
import os
import re
import time
import warnings
from urllib.parse import quote, urlparse, urlsplit, urlunsplit
from pydantic import BaseModel
from fastapi.staticfiles import StaticFiles
import threading
import asyncio

# Global progress tracker
scan_progress: Dict = {
    "is_scanning": False,
    "current_file": "",
    "indexed_count": 0,
    "total_files": 0,
    "progress": 0,
}

# Lock for thread-safe operations
progress_lock = threading.Lock()

# Benchmark V1 state machine - follows scan_progress pattern
# Uses canonical benchmark types from benchmark_types.py
from minirag.api.benchmark_types import (
    BenchmarkState,
    BenchmarkResultItem,
    BenchmarkSummary,
    BenchmarkPageSnapshot,
    BenchmarkModelOption,
    score_to_label,
    is_valid_transition,
    assert_valid_transition,
    build_summary_from_results,
    SCORE_LABELS,
)
from minirag.api.benchmark_adapter import (
    ShipBenchmarkRunnerAdapter,
    compute_recall_rate,
    parse_context_source_ids,
    parse_evidence_source_ids,
    resolve_ship_benchmark_source,
)
from minirag.api.model_registry import (
    LLM_ROLE_BENCHMARK,
    LLM_ROLE_INDEX,
    LLM_ROLE_QUERY,
    RuntimeModelCatalogResponse,
    RuntimeModelSelection,
    build_runtime_model_catalog,
    load_model_registry,
    resolve_runtime_selection,
)
from data_pipeline.source_document_pipeline import (
    build_pipeline_sync_payload,
    derive_doc_dir_name,
    materialize_source_document_to_chunks,
    purge_chunk_outputs_for_doc_name,
    purge_extracted_records_for_doc_name,
)

BENCHMARK_AVAILABLE_MODELS: list[BenchmarkModelOption] = [
    BenchmarkModelOption(id="qwen3.5-2b", label="Qwen 3.5 2B"),
]
BENCHMARK_QUERY_MODES: list[str] = [
    "graph_text_hybrid",
    "graph_only",
    "text_only",
]
BENCHMARK_JUDGE_MODEL = os.getenv(
    "BENCHMARK_JUDGE_MODEL", "deepseek-chat"
)


# Global benchmark state container
# Tracks: run_id, selected_model, lifecycle state, timestamps, progress, summary, results
class BenchmarkStateContainer:
    """
    In-memory benchmark state holder following scan_progress pattern.

    Single active run at a time - second start attempt while active is rejected.
    Terminal snapshots (completed/stopped/failed) are preserved until reset.
    Reset allowed only from terminal states, returns to canonical idle.
    """

    def __init__(self):
        # Core state
        self._state: BenchmarkState = BenchmarkState.idle
        self._run_id: Optional[str] = None
        self._selected_model: str = "qwen3.5-2b"

        # Timestamps
        self._started_at: Optional[datetime] = None
        self._ended_at: Optional[datetime] = None

        # Progress tracking
        self._current_question_index: int = 0
        self._total_questions: int = 0

        # Results and summary
        self._results: list[BenchmarkResultItem] = []
        self._summary: Optional[BenchmarkSummary] = None

        # Error/stop reason
        self._error_message: Optional[str] = None
        self._stopped_reason: Optional[str] = None

        # Active worker lifecycle
        self._worker_thread: Optional[threading.Thread] = None
        self._stop_requested: bool = False
        self._lock = threading.RLock()

        # Terminal snapshot - preserved until reset
        self._terminal_summary: Optional[BenchmarkSummary] = None
        self._terminal_results: list[BenchmarkResultItem] = []
        self._terminal_error_message: Optional[str] = None
        self._terminal_stopped_reason: Optional[str] = None

    # Properties for external access
    @property
    def state(self) -> BenchmarkState:
        return self._state

    @property
    def run_id(self) -> Optional[str]:
        return self._run_id

    @property
    def selected_model(self) -> str:
        return self._selected_model

    @property
    def started_at(self) -> Optional[datetime]:
        return self._started_at

    @property
    def ended_at(self) -> Optional[datetime]:
        return self._ended_at

    @property
    def current_question_index(self) -> int:
        return self._current_question_index

    @property
    def total_questions(self) -> int:
        return self._total_questions

    @property
    def results(self) -> list[BenchmarkResultItem]:
        return self._results

    @property
    def summary(self) -> Optional[BenchmarkSummary]:
        return self._summary

    @property
    def error_message(self) -> Optional[str]:
        return self._error_message

    @property
    def stopped_reason(self) -> Optional[str]:
        return self._stopped_reason

    @property
    def terminal_summary(self) -> Optional[BenchmarkSummary]:
        return self._terminal_summary

    @property
    def terminal_results(self) -> list[BenchmarkResultItem]:
        return self._terminal_results

    @property
    def terminal_error_message(self) -> Optional[str]:
        return self._terminal_error_message

    @property
    def terminal_stopped_reason(self) -> Optional[str]:
        return self._terminal_stopped_reason

    @property
    def progress_percent(self) -> float:
        if self._total_questions == 0:
            return 0.0
        return (self._current_question_index / self._total_questions) * 100

    def is_terminal(self) -> bool:
        """Check if current state is terminal (completed/stopped/failed)."""
        return self._state in [
            BenchmarkState.completed,
            BenchmarkState.stopped,
            BenchmarkState.failed,
        ]

    def can_start(self) -> bool:
        """Check if start is allowed - idle or terminal states."""
        return self._state == BenchmarkState.idle or self.is_terminal()

    def can_stop(self) -> bool:
        """Check if stop is allowed - active states (starting/running)."""
        return self._state in [BenchmarkState.starting, BenchmarkState.running]

    def can_reset(self) -> bool:
        """Check if reset is allowed - terminal states only."""
        return self.is_terminal()

    def set_model(self, model_id: str) -> None:
        """Set the selected model for benchmark runs."""
        self._selected_model = model_id

    def start(self, run_id: str, total_questions: int) -> None:
        """
        Start a new benchmark run.

        Args:
            run_id: Unique identifier for this run
            total_questions: Total number of questions to evaluate

        Raises:
            ValueError: If current state doesn't allow starting (not idle/terminal)
        """
        if not self.can_start():
            raise ValueError(
                f"Cannot start benchmark: current state is {self._state.value}. "
                f"Wait for current run to complete or reset from terminal state."
            )

        # Clear any previous terminal snapshot when starting fresh
        self._clear_terminal_snapshot()

        self._run_id = run_id
        self._state = BenchmarkState.starting
        self._started_at = datetime.now()
        self._ended_at = None
        self._current_question_index = 0
        self._total_questions = total_questions
        self._results = []
        self._summary = None
        self._error_message = None
        self._stopped_reason = None

    def transition_to_running(self) -> None:
        """Transition from starting to running state."""
        if self._state != BenchmarkState.starting:
            raise ValueError(
                f"Cannot transition to running: current state is {self._state.value}"
            )
        self._state = BenchmarkState.running

    def update_progress(self, question_index: int) -> None:
        """Update current question progress."""
        self._current_question_index = question_index

    def add_result(
        self, result: BenchmarkResultItem, question_index: Optional[int] = None
    ) -> None:
        """Add a result item and update summary."""
        with self._lock:
            if question_index is not None:
                self._current_question_index = question_index
            self._results.append(result)
            self._summary = build_summary_from_results(
                self._results, self._total_questions
            )

    def register_worker(self, worker_thread: threading.Thread) -> None:
        """Track the active benchmark worker thread for truthful stop semantics."""
        self._worker_thread = worker_thread

    def clear_worker(self, worker_thread: Optional[threading.Thread] = None) -> None:
        """Clear worker tracking when the active benchmark thread exits."""
        if worker_thread is None or self._worker_thread is worker_thread:
            self._worker_thread = None

    def has_active_worker(self) -> bool:
        """Return whether a tracked worker thread is still alive."""
        return self._worker_thread is not None and self._worker_thread.is_alive()

    def should_stop(self) -> bool:
        """Return whether the current worker should stop after the active question path."""
        return self._stop_requested

    def stop(self, reason: Optional[str] = None) -> None:
        """
        Stop the benchmark run (user-initiated stop).

        Args:
            reason: Optional reason for stopping
        """
        if self._state not in [BenchmarkState.starting, BenchmarkState.running]:
            raise ValueError(
                f"Cannot stop benchmark: current state is {self._state.value}"
            )
        self._state = BenchmarkState.stopping
        self._stopped_reason = reason or "User stopped"
        self._stop_requested = True

    def complete(self) -> None:
        """Mark benchmark as completed normally."""
        if self._state in [BenchmarkState.stopping, BenchmarkState.stopped]:
            self.finalize_stopped()
            return
        if self._state != BenchmarkState.running:
            raise ValueError(
                f"Cannot complete benchmark: current state is {self._state.value}"
            )
        self._state = BenchmarkState.completed
        self._ended_at = datetime.now()
        # Save terminal snapshot
        self._save_terminal_snapshot()

    def fail(self, error_message: str) -> None:
        """
        Mark benchmark as failed with error.

        Args:
            error_message: Error description
        """
        if self._state == BenchmarkState.stopped:
            return
        if self._state not in [BenchmarkState.starting, BenchmarkState.running]:
            raise ValueError(
                f"Cannot fail benchmark: current state is {self._state.value}"
            )
        self._state = BenchmarkState.failed
        self._ended_at = datetime.now()
        self._error_message = error_message
        # Save terminal snapshot
        self._save_terminal_snapshot()

    def finalize_stopped(self) -> None:
        """Finalize stopped state and save terminal snapshot."""
        if self._state == BenchmarkState.stopped:
            return
        if self._state != BenchmarkState.stopping:
            raise ValueError(
                f"Cannot finalize stopped: current state is {self._state.value}"
            )
        self._state = BenchmarkState.stopped
        self._ended_at = datetime.now()
        # Save terminal snapshot
        self._save_terminal_snapshot()

    def finalize_stop_if_worker_finished(self) -> bool:
        """Finalize a pending stop once the worker has really exited."""
        if self._state != BenchmarkState.stopping:
            return False
        if self.has_active_worker():
            return False
        self.finalize_stopped()
        return True

    def worker_finished(self, worker_thread: Optional[threading.Thread] = None) -> None:
        """Clear worker tracking and finalize a pending stop if the worker is gone."""
        self.clear_worker(worker_thread)
        self.finalize_stop_if_worker_finished()

    def reset(self) -> None:
        """
        Reset to canonical idle state.

        Only allowed from terminal states (completed/stopped/failed).
        Clears run_id, timestamps, progress, results, summary, AND terminal snapshot.
        """
        if not self.can_reset():
            raise ValueError(
                f"Cannot reset benchmark: current state is {self._state.value}. "
                f"Reset only allowed from terminal states (completed/stopped/failed)."
            )
        self._state = BenchmarkState.idle
        self._run_id = None
        self._started_at = None
        self._ended_at = None
        self._current_question_index = 0
        self._total_questions = 0
        self._results = []
        self._summary = None
        self._error_message = None
        self._stopped_reason = None
        self._worker_thread = None
        self._stop_requested = False
        # Clear terminal snapshot as part of canonical reset
        self._terminal_summary = None
        self._terminal_results = []
        self._terminal_error_message = None
        self._terminal_stopped_reason = None

    def _save_terminal_snapshot(self) -> None:
        """Save current state as terminal snapshot for UI polling."""
        self._terminal_summary = self._summary
        self._terminal_results = self._results.copy()
        self._terminal_error_message = self._error_message
        self._terminal_stopped_reason = self._stopped_reason
        self._stop_requested = False

    def _clear_terminal_snapshot(self) -> None:
        """Clear terminal snapshot when starting new run."""
        self._terminal_summary = None
        self._terminal_results = []
        self._terminal_error_message = None
        self._terminal_stopped_reason = None

    def to_page_snapshot(self) -> BenchmarkPageSnapshot:
        """
        Convert current state to page snapshot for frontend.

        Uses terminal snapshot if in terminal state (so polling UI can rehydrate).
        """
        # Use terminal snapshot for terminal states, current state otherwise
        with self._lock:
            if self.is_terminal():
                display_summary = self._terminal_summary
                display_results = self._terminal_results[-10:]  # Last 10 results
                display_error = self._terminal_error_message
            else:
                display_summary = self._summary
                display_results = self._results[-10:] if self._results else []
                display_error = self._error_message

            return BenchmarkPageSnapshot(
                state=self._state,
                run_id=self._run_id,
                progress_percent=self.progress_percent,
                summary=display_summary,
                recent_results=display_results,
                available_models=BENCHMARK_AVAILABLE_MODELS,
                selected_model=self._selected_model,
                can_start=self.can_start(),
                can_stop=self.can_stop(),
                can_reset=self.can_reset(),
                error_message=display_error,
            )


# Global benchmark state container instance
benchmark_state = BenchmarkStateContainer()
ship_benchmark_adapter = ShipBenchmarkRunnerAdapter(benchmark_state)

# Backward compatibility - deprecated alias
benchmark_progress: Dict = {}  # Type placeholder only, do not use directly


class IncrementalBenchmarkCsvWriter:
    """Persist benchmark rows incrementally so runs survive process interruption better."""

    def __init__(self, path: str | os.PathLike[str], mode_names: list[str]):
        self.path = Path(path)
        self.mode_names = list(mode_names)
        self._handle = self.path.open("w", encoding="utf-8", newline="")
        self._writer = csv.writer(self._handle)
        headers = [
            "Question_ID",
            "Type",
            "Question",
            "Gold_Answer",
            "Primary_Mode",
            "Primary_Score",
        ]
        for mode in self.mode_names:
            headers.extend([f"{mode}_Answer", f"{mode}_Score", f"{mode}_Recall"])
        self._writer.writerow(headers)
        self._handle.flush()

    def write_item(self, item: BenchmarkResultItem) -> None:
        row = [
            item.question_id,
            item.question_type or "",
            item.question,
            item.gold_answer,
            item.primary_mode or "",
            str(item.raw_score),
        ]
        for mode in self.mode_names:
            row.append(item.mode_answers.get(mode, ""))
            row.append(str(item.mode_scores.get(mode, "")))
            row.append(str(item.mode_recall_rates.get(mode, "")))
        self._writer.writerow(row)
        self._handle.flush()

    def close(self) -> None:
        if not self._handle.closed:
            self._handle.close()


def format_benchmark_exception(exc: Exception) -> str:
    if isinstance(exc, RetryError):
        last_exception = exc.last_attempt.exception()
        if last_exception is not None:
            return format_benchmark_exception(last_exception)
    if isinstance(exc, TimeoutError | asyncio.TimeoutError):
        return "request timed out"
    if isinstance(exc, httpx.HTTPStatusError):
        response = exc.response
        body = ""
        try:
            body = response.text.strip()
        except Exception:
            body = ""
        if body:
            body = body[:500]
            return f"HTTP {response.status_code} from {response.request.url}: {body}"
        return f"HTTP {response.status_code} from {response.request.url}"
    return str(exc)


import json
from datetime import datetime
import xml.etree.ElementTree as ET

from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
import logging
import argparse
import time
import re
from typing import List, Dict, Any, Optional, Union
from minirag import MiniRAG, QueryParam
from minirag.api import __api_version__
from minirag.faultcase_alias_router import (
    FaultcaseAliasStore,
    SUPPORTED_ALIAS_ENTITY_TYPES,
    route_faultcase_query,
)

from minirag.utils import EmbeddingFunc
from minirag.utils import (
    compute_mdhash_id,
    clean_text,
    get_content_summary,
    load_json,
    write_json,
)
from enum import Enum
from pathlib import Path, PureWindowsPath
import shutil
import aiofiles
from ascii_colors import trace_exception, ASCIIColors
from tenacity import RetryError
import httpx
import sys
from uuid import uuid4

from fastapi import Depends, Security
from fastapi.security import APIKeyHeader
from fastapi.middleware.cors import CORSMiddleware
from contextlib import asynccontextmanager

from starlette.status import HTTP_403_FORBIDDEN
import pipmaster as pm
from neo4j import AsyncGraphDatabase

from dotenv import load_dotenv
from minirag.datasource_resolver import (
    DatasourceResolutionError,
    ResolvedDatasource,
    resolve_datasource,
)

load_dotenv()

for stream_name in ("stdout", "stderr"):
    stream = getattr(sys, stream_name, None)
    if stream is not None and hasattr(stream, "reconfigure"):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except OSError:
            pass

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DATASOURCE_ID = "local_ship_docs"


def _normalize_datasource_selector(value: Any) -> Optional[str]:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def get_default_resolved_datasource() -> ResolvedDatasource:
    selector_kwargs = {
        "datasource_id": _normalize_datasource_selector(os.getenv("DATASOURCE_ID")),
        "datasource_root": _normalize_datasource_selector(os.getenv("DATASOURCE_ROOT")),
        "input_dir": _normalize_datasource_selector(os.getenv("INPUT_DIR")),
        "working_dir": _normalize_datasource_selector(os.getenv("WORKING_DIR")),
    }
    if not any(selector_kwargs.values()):
        selector_kwargs["datasource_id"] = DEFAULT_DATASOURCE_ID
    return resolve_datasource(repo_root=REPO_ROOT, **selector_kwargs)


def _resolve_args_datasource(args: argparse.Namespace) -> ResolvedDatasource:
    datasource_id = _normalize_datasource_selector(getattr(args, "datasource_id", None))
    datasource_root = _normalize_datasource_selector(
        getattr(args, "datasource_root", None)
    )
    input_dir = _normalize_datasource_selector(getattr(args, "input_dir", None))
    working_dir = _normalize_datasource_selector(getattr(args, "working_dir", None))

    legacy_input_dir_used = getattr(
        args, "_legacy_input_dir_used", input_dir is not None
    )
    legacy_working_dir_used = getattr(
        args, "_legacy_working_dir_used", working_dir is not None
    )

    if legacy_input_dir_used:
        warnings.warn(
            "--input-dir is deprecated; use --datasource-id or --datasource-root instead.",
            DeprecationWarning,
            stacklevel=3,
        )
    if legacy_working_dir_used:
        warnings.warn(
            "--working-dir is deprecated; use --datasource-id or --datasource-root instead.",
            DeprecationWarning,
            stacklevel=3,
        )

    selector_kwargs = {
        "datasource_id": datasource_id,
        "datasource_root": datasource_root,
        "input_dir": input_dir,
        "working_dir": working_dir,
    }
    if not any(selector_kwargs.values()):
        selector_kwargs["datasource_id"] = DEFAULT_DATASOURCE_ID

    resolved = resolve_datasource(repo_root=REPO_ROOT, **selector_kwargs)
    args._legacy_input_dir_used = legacy_input_dir_used
    args._legacy_working_dir_used = legacy_working_dir_used
    args.datasource = resolved
    args.datasource_id = resolved.id
    args.datasource_root = str(resolved.root)
    args.input_dir = str(resolved.source_root)
    args.source_root = str(resolved.source_root)
    args.staging_root = str(resolved.staging_root)
    args.output_root = str(resolved.output_root)
    args.working_dir = str(resolved.working_dir)
    return resolved


INPUT_DIR_CAPABILITIES: Dict[str, bool] = {
    "mutable": True,
    "deletable": True,
    "reindexable": True,
    "reprocessable": True,
    "read_only": False,
}

PIPELINE_CAPABILITIES: Dict[str, bool] = {
    "mutable": False,
    "deletable": False,
    "reindexable": False,
    "reprocessable": False,
    "read_only": True,
}


def build_resource_descriptor(resource_kind: str, datasource_id: str) -> Dict[str, Any]:
    return {
        "datasource_id": datasource_id,
        "resource_kind": resource_kind,
        "capabilities": dict(
            INPUT_DIR_CAPABILITIES
            if resource_kind == "input_dir"
            else PIPELINE_CAPABILITIES
        ),
    }


def build_datasource_scope(
    resolved_datasource: ResolvedDatasource,
) -> Dict[str, str]:
    return {
        "datasource_id": resolved_datasource.id,
        "datasource_root": str(resolved_datasource.root),
        "source_root": str(resolved_datasource.source_root),
        "staging_root": str(resolved_datasource.staging_root),
        "output_root": str(resolved_datasource.output_root),
        "input_dir": str(resolved_datasource.source_root),
        "working_dir": str(resolved_datasource.working_dir),
    }


def build_faultcase_equipment_catalog(output_root: Path) -> list[dict[str, Any]]:
    exports_root = output_root / "exports" / "cards"
    if not exports_root.exists():
        return []

    equipment_stats: dict[str, dict[str, Any]] = {}

    for overview_path in exports_root.rglob("*_faultcase_cards_overview.txt"):
        try:
            lines = overview_path.read_text(encoding="utf-8").splitlines()
        except OSError:
            continue

        doc_name = ""
        expect_doc_name = False
        for raw_line in lines:
            line = raw_line.strip()
            if not line:
                continue

            if line == "[文档名称]":
                expect_doc_name = True
                continue

            if expect_doc_name:
                doc_name = line
                expect_doc_name = False
                continue

            if not line.startswith("- ") or "->" not in line:
                continue

            _, _, content = line.partition(":")
            equipment_name, _, _fault_name = content.partition("->")
            equipment = equipment_name.strip()
            if not equipment:
                continue

            stat = equipment_stats.setdefault(
                equipment,
                {
                    "name": equipment,
                    "faultcase_count": 0,
                    "documents": set(),
                },
            )
            stat["faultcase_count"] += 1
            if doc_name:
                stat["documents"].add(doc_name)

    items = []
    for payload in equipment_stats.values():
        documents = sorted(payload["documents"])
        items.append(
            {
                "name": payload["name"],
                "faultcase_count": payload["faultcase_count"],
                "document_count": len(documents),
                "documents": documents[:3],
            }
        )

    items.sort(key=lambda item: (-item["faultcase_count"], item["name"]))
    return items


def _pipeline_tombstones_file(datasource: ResolvedDatasource) -> Path:
    return datasource.staging_root / ".deleted_pipeline_docs.json"


def _load_pipeline_tombstones(datasource: ResolvedDatasource) -> set[str]:
    payload = load_json(str(_pipeline_tombstones_file(datasource)))
    if isinstance(payload, list):
        return {str(item) for item in payload}
    if isinstance(payload, dict):
        return {str(key) for key, value in payload.items() if value}
    return set()


def _save_pipeline_tombstones(
    datasource: ResolvedDatasource, tombstones: set[str]
) -> None:
    tombstones_file = _pipeline_tombstones_file(datasource)
    tombstones_file.parent.mkdir(parents=True, exist_ok=True)
    write_json(sorted(tombstones), str(tombstones_file))


def tombstone_pipeline_output(name: str, datasource: ResolvedDatasource) -> None:
    tombstones = _load_pipeline_tombstones(datasource)
    tombstones.add(name)
    _save_pipeline_tombstones(datasource, tombstones)


def revive_pipeline_output(name: str, datasource: ResolvedDatasource) -> None:
    tombstones = _load_pipeline_tombstones(datasource)
    if name not in tombstones:
        return
    tombstones.discard(name)
    _save_pipeline_tombstones(datasource, tombstones)


def _delete_tree(path: Path) -> None:
    if not path.exists():
        return
    if path.is_file() or path.is_symlink():
        path.unlink(missing_ok=True)
        return
    for child in sorted(path.iterdir(), reverse=True):
        _delete_tree(child)
    path.rmdir()


def purge_pipeline_outputs_for_source_file(
    file_path: Path, datasource: ResolvedDatasource
) -> bool:
    source_name = Path(file_path).name
    source_stem = Path(file_path).stem
    doc_dir_name = derive_doc_dir_name(Path(file_path))
    deleted_any = False
    exact_matches: list[Path] = []
    records_root = datasource.staging_root / "extracted" / "records"

    if records_root.exists():
        for records_path in sorted(records_root.rglob("diagnostic_records_llm.json")):
            if not records_path.exists():
                continue

            try:
                payload = load_json(str(records_path)) or {}
            except Exception:
                payload = {}

            payload_doc_name = str(payload.get("doc_name") or "")
            record_owner = records_path.parent.name
            if record_owner in {source_stem, source_name} and payload_doc_name in {
                source_stem,
                source_name,
                "",
            }:
                exact_matches.append(records_path.parent)

    if len(exact_matches) == 1:
        _delete_tree(exact_matches[0])
        deleted_any = True

    if purge_chunk_outputs_for_doc_name(datasource, doc_dir_name):
        deleted_any = True

    if purge_extracted_records_for_doc_name(datasource, doc_dir_name):
        deleted_any = True

    tombstone_pipeline_output(source_stem, datasource)
    tombstone_pipeline_output(source_name, datasource)
    return deleted_any


def get_default_working_dir() -> str:
    env_override = os.getenv("WORKING_DIR")
    if env_override:
        return env_override
    return str(get_default_resolved_datasource().working_dir)


def get_default_input_dir() -> str:
    return os.getenv("INPUT_DIR") or str(get_default_resolved_datasource().source_root)


def get_default_embedding_dim() -> int:
    env_override = os.getenv("EMBEDDING_DIM")
    if env_override:
        try:
            return int(env_override)
        except ValueError:
            pass

    working_dir = Path(get_default_working_dir())
    for vector_store_name in (
        "vdb_entities.json",
        "vdb_relationships.json",
        "vdb_chunks.json",
    ):
        vector_store_path = working_dir / vector_store_name
        if not vector_store_path.exists():
            continue
        try:
            payload = json.loads(vector_store_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        embedding_dim = payload.get("embedding_dim")
        if isinstance(embedding_dim, int) and embedding_dim > 0:
            return embedding_dim

    return 1024


def estimate_tokens(text: str) -> int:
    """Estimate the number of tokens in text
    Chinese characters: approximately 1.5 tokens per character
    English characters: approximately 0.25 tokens per character
    """
    # Use regex to match Chinese and non-Chinese characters separately
    chinese_chars = len(re.findall(r"[\u4e00-\u9fff]", text))
    non_chinese_chars = len(re.findall(r"[^\u4e00-\u9fff]", text))

    # Calculate estimated token count
    tokens = chinese_chars * 1.5 + non_chinese_chars * 0.25

    return int(tokens)


class OllamaServerInfos:
    # Constants for emulated Ollama model information
    LIGHTRAG_NAME = "minirag"
    LIGHTRAG_TAG = os.getenv("OLLAMA_EMULATING_MODEL_TAG", "latest")
    LIGHTRAG_MODEL = f"{LIGHTRAG_NAME}:{LIGHTRAG_TAG}"
    LIGHTRAG_SIZE = 7365960935  # it's a dummy value
    LIGHTRAG_CREATED_AT = "2024-01-15T00:00:00Z"
    LIGHTRAG_DIGEST = "sha256:minirag"

    KV_STORAGE = "JsonKVStorage"
    DOC_STATUS_STORAGE = "JsonDocStatusStorage"
    GRAPH_STORAGE = "NetworkXStorage"
    VECTOR_STORAGE = "NanoVectorDBStorage"


# Add infos
ollama_server_infos = OllamaServerInfos()

LOCAL_MODEL_HOST = "http://127.0.0.1:1234/v1"
OLLAMA_DEFAULT_HOST = "http://127.0.0.1:11434"
LOLLMS_DEFAULT_HOST = "http://127.0.0.1:9600"


def get_default_host(binding_type: str) -> str:
    normalized = str(binding_type or "").strip().lower()
    if normalized == "ollama":
        return os.getenv("OLLAMA_BASE_URL", OLLAMA_DEFAULT_HOST)
    if normalized == "openai-ollama":
        return os.getenv("OLLAMA_BASE_URL", f"{OLLAMA_DEFAULT_HOST}/v1")
    if normalized in {"openai", "deepseek"}:
        return os.getenv("OPENAI_BASE_URL", os.getenv("LMSTUDIO_HOST", LOCAL_MODEL_HOST))
    if normalized == "azure_openai":
        return os.getenv("AZURE_OPENAI_ENDPOINT", LOCAL_MODEL_HOST)
    if normalized == "lollms":
        return os.getenv("LOLLMS_BASE_URL", LOLLMS_DEFAULT_HOST)
    return os.getenv("LLM_BINDING_HOST", LOCAL_MODEL_HOST)


class RuntimeUnloadResult(BaseModel):
    status: str
    binding: str | None = None
    host: str | None = None
    model: str | None = None
    instance_ids: list[str] = []
    message: str | None = None


def _normalize_lmstudio_rest_base(host: str | None) -> str | None:
    if not host:
        return None

    normalized = host.rstrip("/")
    parts = urlsplit(normalized)
    path = parts.path.rstrip("/")
    if path.endswith("/v1"):
        path = path[: -len("/v1")] + "/api/v1"
    elif not path.endswith("/api/v1"):
        path = f"{path}/api/v1" if path else "/api/v1"
    return urlunsplit((parts.scheme, parts.netloc, path, parts.query, parts.fragment))


def _build_lmstudio_admin_headers(api_key: str | None) -> dict[str, str]:
    normalized_key = (api_key or "").strip()
    headers = {"Content-Type": "application/json"}
    if normalized_key and normalized_key.lower() not in {
        "empty",
        "lm-studio",
        "ollama",
    }:
        headers["Authorization"] = f"Bearer {normalized_key}"
    return headers


def _physical_model_signature(config) -> tuple[str, str | None, str]:
    return (
        str(getattr(config, "binding", "") or "").strip().lower(),
        _normalize_lmstudio_rest_base(getattr(config, "host", None)),
        str(getattr(config, "model", "") or "").strip(),
    )


async def unload_lmstudio_model_instances(
    *,
    binding: str,
    host: str | None,
    model: str,
    api_key: str | None = None,
) -> RuntimeUnloadResult:
    normalized_binding = str(binding or "").strip().lower()
    if normalized_binding not in {"openai", "deepseek"}:
        return RuntimeUnloadResult(
            status="skipped",
            binding=binding,
            host=host,
            model=model,
            message="binding is not routed through LM Studio REST management",
        )

    base_url = _normalize_lmstudio_rest_base(host)
    if not base_url:
        return RuntimeUnloadResult(
            status="skipped",
            binding=binding,
            host=host,
            model=model,
            message="host is not configured",
        )

    parsed = urlparse(base_url)
    trust_env = (parsed.hostname or "").strip().lower() not in {
        "127.0.0.1",
        "localhost",
        "::1",
    }
    headers = _build_lmstudio_admin_headers(api_key)

    try:
        async with httpx.AsyncClient(timeout=10, trust_env=trust_env) as client:
            models_response = await client.get(f"{base_url}/models", headers=headers)
            models_response.raise_for_status()
            payload = models_response.json()

            instance_ids: list[str] = []
            for item in payload.get("models", []):
                if item.get("key") != model:
                    continue
                for loaded in item.get("loaded_instances", []):
                    instance_id = str(loaded.get("id", "")).strip()
                    if instance_id:
                        instance_ids.append(instance_id)

            if not instance_ids:
                return RuntimeUnloadResult(
                    status="not_loaded",
                    binding=binding,
                    host=host,
                    model=model,
                    message="no loaded instances found",
                )

            for instance_id in instance_ids:
                unload_response = await client.post(
                    f"{base_url}/models/unload",
                    headers=headers,
                    json={"instance_id": instance_id},
                )
                unload_response.raise_for_status()

            return RuntimeUnloadResult(
                status="unloaded",
                binding=binding,
                host=host,
                model=model,
                instance_ids=instance_ids,
            )
    except Exception as exc:
        return RuntimeUnloadResult(
            status="failed",
            binding=binding,
            host=host,
            model=model,
            message=str(exc),
        )


def get_env_value(env_key: str, default: Any, value_type: type = str) -> Any:
    """
    Get value from environment variable with type conversion

    Args:
        env_key (str): Environment variable key
        default (Any): Default value if env variable is not set
        value_type (type): Type to convert the value to

    Returns:
        Any: Converted value from environment or default
    """
    value = os.getenv(env_key)
    if value is None:
        return default

    if isinstance(value_type, bool):
        return value.lower() in ("true", "1", "yes")
    try:
        return value_type(value)
    except ValueError:
        return default


async def _verify_neo4j_connection(args: argparse.Namespace) -> None:
    driver = AsyncGraphDatabase.driver(
        args.neo4j_uri,
        auth=(args.neo4j_username, args.neo4j_password),
    )
    try:
        async with driver.session(database=args.neo4j_database) as session:
            result = await session.run("RETURN 1 AS ok")
            record = await result.single()
            if not record or record["ok"] != 1:
                raise Exception(
                    "Neo4j connectivity check returned an unexpected result"
                )
    finally:
        await driver.close()


def display_splash_screen(args: argparse.Namespace) -> None:
    """
    Display a colorful splash screen showing MiniRAG server configuration

    Args:
        args: Parsed command line arguments
    """
    # Banner
    ASCIIColors.cyan(f"""
    ╔══════════════════════════════════════════════════════════════╗
    ║                   🚀 MiniRAG Server v{__api_version__}                  ║
    ║          Fast, Lightweight RAG Server Implementation         ║
    ╚══════════════════════════════════════════════════════════════╝
    """)

    # Server Configuration
    ASCIIColors.magenta("\n📡 Server Configuration:")
    ASCIIColors.white("    ├─ Host: ", end="")
    ASCIIColors.yellow(f"{args.host}")
    ASCIIColors.white("    ├─ Port: ", end="")
    ASCIIColors.yellow(f"{args.port}")
    ASCIIColors.white("    ├─ SSL Enabled: ", end="")
    ASCIIColors.yellow(f"{args.ssl}")
    if args.ssl:
        ASCIIColors.white("    ├─ SSL Cert: ", end="")
        ASCIIColors.yellow(f"{args.ssl_certfile}")
        ASCIIColors.white("    └─ SSL Key: ", end="")
        ASCIIColors.yellow(f"{args.ssl_keyfile}")

    # Directory Configuration
    ASCIIColors.magenta("\n📂 Directory Configuration:")
    ASCIIColors.white("    ├─ Working Directory: ", end="")
    ASCIIColors.yellow(f"{args.working_dir}")
    ASCIIColors.white("    └─ Input Directory: ", end="")
    ASCIIColors.yellow(f"{args.input_dir}")

    # LLM Configuration
    ASCIIColors.magenta("\n🤖 LLM Configuration:")
    ASCIIColors.white("    ├─ Binding: ", end="")
    ASCIIColors.yellow(f"{args.llm_binding}")
    ASCIIColors.white("    ├─ Host: ", end="")
    ASCIIColors.yellow(f"{args.llm_binding_host}")
    ASCIIColors.white("    └─ Model: ", end="")
    ASCIIColors.yellow(f"{args.llm_model}")

    # Embedding Configuration
    ASCIIColors.magenta("\n📊 Embedding Configuration:")
    ASCIIColors.white("    ├─ Binding: ", end="")
    ASCIIColors.yellow(f"{args.embedding_binding}")
    ASCIIColors.white("    ├─ Host: ", end="")
    ASCIIColors.yellow(f"{args.embedding_binding_host}")
    ASCIIColors.white("    ├─ Model: ", end="")
    ASCIIColors.yellow(f"{args.embedding_model}")
    ASCIIColors.white("    └─ Dimensions: ", end="")
    ASCIIColors.yellow(f"{args.embedding_dim}")

    # RAG Configuration
    ASCIIColors.magenta("\n⚙️ RAG Configuration:")
    ASCIIColors.white("    ├─ Max Async Operations: ", end="")
    ASCIIColors.yellow(f"{args.max_async}")
    ASCIIColors.white("    ├─ Max Tokens: ", end="")
    ASCIIColors.yellow(f"{args.max_tokens}")
    ASCIIColors.white("    ├─ Max Embed Tokens: ", end="")
    ASCIIColors.yellow(f"{args.max_embed_tokens}")
    ASCIIColors.white("    ├─ Chunk Size: ", end="")
    ASCIIColors.yellow(f"{args.chunk_size}")
    ASCIIColors.white("    ├─ Chunk Overlap Size: ", end="")
    ASCIIColors.yellow(f"{args.chunk_overlap_size}")
    ASCIIColors.white("    ├─ History Turns: ", end="")
    ASCIIColors.yellow(f"{args.history_turns}")
    ASCIIColors.white("    ├─ Cosine Threshold: ", end="")
    ASCIIColors.yellow(f"{args.cosine_threshold}")
    ASCIIColors.white("    └─ Top-K: ", end="")
    ASCIIColors.yellow(f"{args.top_k}")

    # System Configuration
    ASCIIColors.magenta("\n🛠️ System Configuration:")
    ASCIIColors.white("    ├─ Ollama Emulating Model: ", end="")
    ASCIIColors.yellow(f"{ollama_server_infos.LIGHTRAG_MODEL}")
    ASCIIColors.white("    ├─ Log Level: ", end="")
    ASCIIColors.yellow(f"{args.log_level}")
    ASCIIColors.white("    ├─ Timeout: ", end="")
    ASCIIColors.yellow(f"{args.timeout if args.timeout else 'None (infinite)'}")
    ASCIIColors.white("    └─ API Key: ", end="")
    ASCIIColors.yellow("Set" if args.key else "Not Set")

    # Server Status
    ASCIIColors.green("\n✨ Server starting up...\n")

    # Server Access Information
    protocol = "https" if args.ssl else "http"
    if args.host == "0.0.0.0":
        ASCIIColors.magenta("\n🌐 Server Access Information:")
        ASCIIColors.white("    ├─ Local Access: ", end="")
        ASCIIColors.yellow(f"{protocol}://localhost:{args.port}")
        ASCIIColors.white("    ├─ Remote Access: ", end="")
        ASCIIColors.yellow(f"{protocol}://<your-ip-address>:{args.port}")
        ASCIIColors.white("    ├─ API Documentation (local): ", end="")
        ASCIIColors.yellow(f"{protocol}://localhost:{args.port}/docs")
        ASCIIColors.white("    └─ Alternative Documentation (local): ", end="")
        ASCIIColors.yellow(f"{protocol}://localhost:{args.port}/redoc")

        ASCIIColors.yellow("\n📝 Note:")
        ASCIIColors.white("""    Since the server is running on 0.0.0.0:
    - Use 'localhost' or '127.0.0.1' for local access
    - Use your machine's IP address for remote access
    - To find your IP address:
      • Windows: Run 'ipconfig' in terminal
      • Linux/Mac: Run 'ifconfig' or 'ip addr' in terminal
    """)
    else:
        base_url = f"{protocol}://{args.host}:{args.port}"
        ASCIIColors.magenta("\n🌐 Server Access Information:")
        ASCIIColors.white("    ├─ Base URL: ", end="")
        ASCIIColors.yellow(f"{base_url}")
        ASCIIColors.white("    ├─ API Documentation: ", end="")
        ASCIIColors.yellow(f"{base_url}/docs")
        ASCIIColors.white("    └─ Alternative Documentation: ", end="")
        ASCIIColors.yellow(f"{base_url}/redoc")

    # Usage Examples
    ASCIIColors.magenta("\n📚 Quick Start Guide:")
    ASCIIColors.cyan("""
    1. Access the Swagger UI:
       Open your browser and navigate to the API documentation URL above

    2. API Authentication:""")
    if args.key:
        ASCIIColors.cyan("""       Add the following header to your requests:
       X-API-Key: <your-api-key>
    """)
    else:
        ASCIIColors.cyan("       No authentication required\n")

    ASCIIColors.cyan("""    3. Basic Operations:
       - POST /upload_document: Upload new documents to RAG
       - POST /query: Query your document collection
       - GET /collections: List available collections

    4. Monitor the server:
       - Check server logs for detailed operation information
       - Use healthcheck endpoint: GET /health
    """)

    # Security Notice
    if args.key:
        ASCIIColors.yellow("\n⚠️  Security Notice:")
        ASCIIColors.white("""    API Key authentication is enabled.
    Make sure to include the X-API-Key header in all your requests.
    """)

    ASCIIColors.green("Server is ready to accept connections! 🚀\n")

    # Ensure splash output flush to system log
    sys.stdout.flush()


def parse_args() -> argparse.Namespace:
    """
    Parse command line arguments with environment variable fallback

    Returns:
        argparse.Namespace: Parsed arguments
    """

    parser = argparse.ArgumentParser(
        description="MiniRAG FastAPI Server with separate working and input directories"
    )

    def bool_env_type(value):
        if isinstance(value, bool):
            return value
        normalized = str(value).strip().lower()
        return normalized in {"1", "true", "yes", "on"}

    # Bindings configuration
    parser.add_argument(
        "--llm-binding",
        default=get_env_value("LLM_BINDING", "ollama"),
        help="LLM binding to be used. Current mainline only supports: ollama",
    )
    parser.add_argument(
        "--embedding-binding",
        default=get_env_value("EMBEDDING_BINDING", "ollama"),
        help="Embedding binding to be used. Current mainline only supports: ollama",
    )

    # Server configuration
    parser.add_argument(
        "--host",
        default=get_env_value("HOST", "0.0.0.0"),
        help="Server host (default: from env or 0.0.0.0)",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=get_env_value("PORT", 9733, int),
        help="Server port (default: from env or 9733)",
    )
    parser.add_argument(
        "--serve-static-ui",
        dest="serve_static_ui",
        action="store_true",
        default=get_env_value("SERVE_STATIC_UI", True, bool_env_type),
        help="Mount the built frontend static assets at '/'. Enabled by default for web mode.",
    )
    parser.add_argument(
        "--no-static-ui",
        dest="serve_static_ui",
        action="store_false",
        help="Disable frontend static asset mounting so the server runs as API-only backend for desktop shell mode.",
    )

    # Directory configuration
    parser.add_argument(
        "--working-dir",
        default=os.getenv("WORKING_DIR"),
        help="DEPRECATED compatibility selector for the runtime working directory; use --datasource-id or --datasource-root instead.",
    )
    parser.add_argument(
        "--input-dir",
        default=os.getenv("INPUT_DIR"),
        help="DEPRECATED compatibility selector for the datasource source root; use --datasource-id or --datasource-root instead.",
    )
    parser.add_argument(
        "--datasource-id",
        default=get_env_value("DATASOURCE_ID", ""),
        help="Preferred datasource identifier for graph/doc routes.",
    )
    parser.add_argument(
        "--datasource-root",
        default=get_env_value("DATASOURCE_ROOT", None),
        help="Explicit datasource root path for graph/doc routes.",
    )
    parser.add_argument(
        "--model-registry",
        default=get_env_value(
            "MODEL_REGISTRY_PATH", str(REPO_ROOT / "config" / "model_registry.json")
        ),
        help="JSON model registry path for runtime model identifiers.",
    )
    parser.add_argument(
        "--graph-storage",
        default=get_env_value("GRAPH_STORAGE", "NetworkXStorage"),
        choices=["NetworkXStorage", "Neo4JStorage"],
        help="Graph backend for the current mainline",
    )
    parser.add_argument(
        "--graph-storage-fallback",
        default=get_env_value("GRAPH_STORAGE_FALLBACK", "none"),
        choices=["none", "NetworkXStorage"],
        help="Fallback graph backend if the configured graph backend is unavailable",
    )
    parser.add_argument(
        "--neo4j-uri",
        default=get_env_value("NEO4J_URI", None),
        help="Neo4j bolt URI when using Neo4JStorage",
    )
    parser.add_argument(
        "--neo4j-username",
        default=get_env_value("NEO4J_USERNAME", os.getenv("NEO4J_USER")),
        help="Neo4j username when using Neo4JStorage",
    )
    parser.add_argument(
        "--neo4j-password",
        default=get_env_value("NEO4J_PASSWORD", None),
        help="Neo4j password when using Neo4JStorage",
    )
    parser.add_argument(
        "--neo4j-database",
        default=get_env_value("NEO4J_DATABASE", "neo4j"),
        help="Neo4j database name when using Neo4JStorage",
    )

    # LLM Model configuration
    parser.add_argument(
        "--llm-binding-host",
        default=get_env_value("LLM_BINDING_HOST", None),
        help="LLM server host URL. If not provided, defaults to http://localhost:11434",
    )

    default_llm_api_key = get_env_value(
        "LLM_BINDING_API_KEY",
        os.getenv("DEEPSEEK_API_KEY"),
    )

    parser.add_argument(
        "--llm-binding-api-key",
        default=default_llm_api_key,
        help="llm server API key (default: from env or empty string)",
    )

    parser.add_argument(
        "--llm-model",
        default=get_env_value("LLM_MODEL", "qwen3.5-2b"),
        help="LLM model name (default: from env or qwen3.5-2b)",
    )
    parser.add_argument(
        "--query-llm-binding",
        default=get_env_value("QUERY_LLM_BINDING", "ollama"),
        help="Query-time LLM binding. Current mainline only supports: ollama",
    )
    parser.add_argument(
        "--query-llm-binding-host",
        default=get_env_value("QUERY_LLM_BINDING_HOST", None),
        help="Query-time LLM host URL. If not provided, defaults based on query-llm-binding.",
    )
    default_query_llm_api_key = get_env_value(
        "QUERY_LLM_BINDING_API_KEY",
        ""
        if get_env_value("QUERY_LLM_BINDING", "ollama") == "ollama"
        else get_env_value("LLM_BINDING_API_KEY", os.getenv("DEEPSEEK_API_KEY")),
    )
    parser.add_argument(
        "--query-llm-binding-api-key",
        default=default_query_llm_api_key,
        help="Query-time llm server API key",
    )
    parser.add_argument(
        "--query-llm-model",
        default=get_env_value("QUERY_LLM_MODEL", "qwen3.5-2b"),
        help="Query-time LLM model name (default: QUERY_LLM_MODEL, then qwen3.5-2b)",
    )

    parser.add_argument(
        "--batch-insert-size",
        type=int,
        default=get_env_value("BATCH_INSERT_SIZE", 10, int),
        help="Number of files to batch into a single insert call during directory scans (default: from env or 10)",
    )
    parser.add_argument(
        "--max-parallel-insert",
        type=int,
        default=get_env_value("MAX_PARALLEL_INSERT", 10, int),
        help="Maximum number of documents MiniRAG processes concurrently inside a batch insert (default: from env or 10)",
    )

    # Embedding model configuration
    parser.add_argument(
        "--embedding-binding-host",
        default=get_env_value("EMBEDDING_BINDING_HOST", None),
        help="Embedding server host URL. If not provided, defaults to http://localhost:11434",
    )

    default_embedding_api_key = get_env_value("EMBEDDING_BINDING_API_KEY", "")
    parser.add_argument(
        "--embedding-binding-api-key",
        default=default_embedding_api_key,
        help="embedding server API key (default: from env or empty string)",
    )

    parser.add_argument(
        "--embedding-model",
        default=get_env_value("EMBEDDING_MODEL", "text-embedding-qwen3-embedding-4b"),
        help="Embedding model name (default: from env or text-embedding-qwen3-embedding-4b)",
    )

    parser.add_argument(
        "--chunk_size",
        default=get_env_value("CHUNK_SIZE", 1200),
        help="chunk chunk size default 1200",
    )

    parser.add_argument(
        "--chunk_overlap_size",
        default=get_env_value("CHUNK_OVERLAP_SIZE", 100),
        help="chunk overlap size default 100",
    )

    def timeout_type(value):
        if value is None or value == "None":
            return None
        return int(value)

    parser.add_argument(
        "--timeout",
        default=get_env_value("TIMEOUT", None, timeout_type),
        type=timeout_type,
        help="Timeout in seconds (useful when using slow AI). Use None for infinite timeout",
    )

    # RAG configuration
    parser.add_argument(
        "--max-async",
        type=int,
        default=get_env_value("MAX_ASYNC", 4, int),
        help="Maximum async operations (default: from env or 4)",
    )
    parser.add_argument(
        "--max-tokens",
        type=int,
        default=get_env_value("MAX_TOKENS", 32768, int),
        help="Maximum token size (default: from env or 32768)",
    )
    parser.add_argument(
        "--embedding-dim",
        type=int,
        default=get_default_embedding_dim(),
        help="Embedding dimensions (default: EMBEDDING_DIM, then infer from selected workdir, else 1024)",
    )
    parser.add_argument(
        "--max-embed-tokens",
        type=int,
        default=get_env_value("MAX_EMBED_TOKENS", 8192, int),
        help="Maximum embedding token size (default: from env or 8192)",
    )

    # Logging configuration
    parser.add_argument(
        "--log-level",
        default=get_env_value("LOG_LEVEL", "INFO"),
        choices=["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"],
        help="Logging level (default: from env or INFO)",
    )

    parser.add_argument(
        "--key",
        type=str,
        default=get_env_value("LIGHTRAG_API_KEY", None),
        help="API key for authentication. This protects minirag server against unauthorized access",
    )

    # Optional https parameters
    parser.add_argument(
        "--ssl",
        action="store_true",
        default=get_env_value("SSL", False, bool),
        help="Enable HTTPS (default: from env or False)",
    )
    parser.add_argument(
        "--ssl-certfile",
        default=get_env_value("SSL_CERTFILE", None),
        help="Path to SSL certificate file (required if --ssl is enabled)",
    )
    parser.add_argument(
        "--ssl-keyfile",
        default=get_env_value("SSL_KEYFILE", None),
        help="Path to SSL private key file (required if --ssl is enabled)",
    )
    parser.add_argument(
        "--auto-scan-at-startup",
        action="store_true",
        default=False,
        help="Enable automatic scanning when the program starts",
    )

    parser.add_argument(
        "--history-turns",
        type=int,
        default=get_env_value("HISTORY_TURNS", 3, int),
        help="Number of conversation history turns to include (default: from env or 3)",
    )

    # Search parameters
    parser.add_argument(
        "--top-k",
        type=int,
        default=get_env_value("TOP_K", 50, int),
        help="Number of most similar results to return (default: from env or 50)",
    )
    parser.add_argument(
        "--cosine-threshold",
        type=float,
        default=get_env_value("COSINE_THRESHOLD", 0.4, float),
        help="Cosine similarity threshold (default: from env or 0.4)",
    )

    parser.add_argument(
        "--simulated-model-name",
        type=str,
        default=get_env_value(
            "SIMULATED_MODEL_NAME", ollama_server_infos.LIGHTRAG_MODEL
        ),
        help="Number of conversation history turns to include (default: from env or 3)",
    )

    args = parser.parse_args()

    ollama_server_infos.LIGHTRAG_MODEL = args.simulated_model_name

    try:
        _resolve_args_datasource(args)
    except DatasourceResolutionError as exc:
        parser.error(str(exc))

    return args


class DocumentManager:
    """Handles document operations and tracking"""

    def __init__(
        self,
        input_dir: str,
        supported_extensions: tuple = (".txt", ".md", ".xlsx"),
    ):
        self.input_dir = Path(input_dir)
        self.supported_extensions = supported_extensions
        self.indexed_files = set()

        # Create input directory if it doesn't exist
        self.input_dir.mkdir(parents=True, exist_ok=True)

    def scan_directory_for_new_files(self) -> List[Path]:
        """Scan input directory for new files"""
        new_files = []
        for ext in self.supported_extensions:
            for file_path in self.input_dir.rglob(f"*{ext}"):
                if file_path not in self.indexed_files:
                    new_files.append(file_path)
        return new_files

    def scan_directory(self) -> List[Path]:
        """Scan input directory for new files"""
        new_files = []
        for ext in self.supported_extensions:
            for file_path in self.input_dir.rglob(f"*{ext}"):
                new_files.append(file_path)
        return new_files

    def mark_as_indexed(self, file_path: Path):
        """Mark a file as indexed"""
        self.indexed_files.add(file_path)

    def is_supported_file(self, filename: str) -> bool:
        """Check if file type is supported"""
        return any(filename.lower().endswith(ext) for ext in self.supported_extensions)

    def resolve_upload_target(self, filename: str | None) -> Path:
        """Resolve a safe upload filename under input_dir."""
        raw_name = (filename or "").strip()
        safe_name = PureWindowsPath(raw_name).name
        if (
            not raw_name
            or safe_name != raw_name
            or safe_name in {".", ".."}
            or "/" in raw_name
            or "\\" in raw_name
            or ":" in raw_name
        ):
            raise HTTPException(status_code=400, detail="Invalid upload filename")
        if not self.is_supported_file(safe_name):
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported file type. Supported types: {self.supported_extensions}",
            )
        target = (self.input_dir / safe_name).resolve()
        try:
            target.relative_to(self.input_dir.resolve())
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid upload path") from exc
        return target


class DocumentRegistry:
    """Persist file-centric indexing metadata for the frontend."""

    def __init__(self, working_dir: str, input_dir: str, datasource_id: str):
        self.input_dir = Path(input_dir).resolve()
        self.datasource_id = datasource_id.strip()
        self.store_file = Path(working_dir) / "document_registry.json"
        self._data: Dict[str, Dict[str, Any]] = load_json(str(self.store_file)) or {}

    def sync_from_disk(self, supported_extensions: tuple[str, ...]) -> None:
        existing_keys = set()
        for ext in supported_extensions:
            for file_path in self.input_dir.rglob(f"*{ext}"):
                key = self._key_for(file_path)
                existing_keys.add(key)
                current = self._data.get(key, {})
                self._data[key] = {
                    **current,
                    **self._snapshot_for(file_path),
                }

        stale_keys = [key for key in self._data.keys() if key not in existing_keys]
        for key in stale_keys:
            self._data.pop(key, None)

        self._save()

    def mark_processing(self, file_path: Path, doc_id: str) -> None:
        self._update(
            file_path,
            {
                "doc_id": doc_id,
                "status": "processing",
                "error": None,
                "last_event_at": datetime.now().isoformat(),
            },
        )

    def mark_indexed(
        self, file_path: Path, doc_id: str, doc_status: Optional[Dict[str, Any]]
    ) -> None:
        payload: Dict[str, Any] = {
            "doc_id": doc_id,
            "status": "indexed",
            "error": None,
            "indexed_at": datetime.now().isoformat(),
            "last_event_at": datetime.now().isoformat(),
        }
        if doc_status:
            payload.update(
                {
                    "chunks_count": doc_status.get("chunks_count"),
                    "content_length": doc_status.get("content_length"),
                    "content_summary": doc_status.get("content_summary"),
                    "doc_status": doc_status.get("status"),
                    "updated_at": doc_status.get("updated_at"),
                }
            )
        self._update(file_path, payload)

    def mark_failed(
        self, file_path: Path, error: str, doc_id: Optional[str] = None
    ) -> None:
        payload: Dict[str, Any] = {
            "status": "failed",
            "error": error,
            "last_event_at": datetime.now().isoformat(),
        }
        if doc_id:
            payload["doc_id"] = doc_id
        self._update(file_path, payload)

    def get_indexed_paths(self) -> set[Path]:
        indexed_paths: set[Path] = set()
        for item in self._data.values():
            if item.get("status") != "indexed":
                continue
            absolute_path = item.get("absolute_path")
            if absolute_path and Path(absolute_path).exists():
                indexed_paths.add(Path(absolute_path))
        return indexed_paths

    def build_summary(self) -> Dict[str, Any]:
        items = sorted(
            self._data.values(),
            key=lambda item: item.get("modified_at", ""),
            reverse=True,
        )
        stats = {
            "total": len(items),
            "indexed": 0,
            "processing": 0,
            "failed": 0,
            "pending": 0,
        }
        for item in items:
            status = item.get("status", "pending")
            if status not in stats:
                stats["pending"] += 1
            else:
                stats[status] += 1

        return {"datasource_id": self.datasource_id, "stats": stats, "items": items}

    def _update(self, file_path: Path, payload: Dict[str, Any]) -> None:
        key = self._key_for(file_path)
        current = self._data.get(key, {})
        self._data[key] = {
            **current,
            **self._snapshot_for(file_path),
            **payload,
        }
        self._save()

    def _snapshot_for(self, file_path: Path) -> Dict[str, Any]:
        file_path = Path(file_path).resolve()
        stats = file_path.stat()
        return {
            "datasource_id": self.datasource_id,
            "name": file_path.name,
            "relative_path": self._key_for(file_path),
            "absolute_path": str(file_path),
            "type": file_path.suffix.lstrip(".").upper(),
            "size": stats.st_size,
            "modified_at": datetime.fromtimestamp(stats.st_mtime).isoformat(),
            **build_resource_descriptor("input_dir", self.datasource_id),
        }

    def _key_for(self, file_path: Path) -> str:
        file_path = Path(file_path).resolve()
        try:
            return str(file_path.relative_to(self.input_dir))
        except ValueError:
            return file_path.name

    def _save(self) -> None:
        self.store_file.parent.mkdir(parents=True, exist_ok=True)
        write_json(self._data, str(self.store_file))

    def delete(self, file_path: Path) -> None:
        key = self._key_for(file_path)
        self._data.pop(key, None)
        self._save()


class ChatSessionStore:
    """Persist datasource-scoped chat sessions for the offline demo runtime."""

    _ALLOWED_MESSAGE_KEYS = {
        "id",
        "role",
        "content",
        "timestamp",
        "query",
        "mode",
        "endpoint",
        "latencyMs",
        "datasourceId",
        "evidenceStatus",
        "evidenceItems",
        "contextSources",
        "contextEntities",
        "contextRaw",
        "evidenceNote",
        "streaming",
        "error",
    }

    def __init__(
        self,
        working_dir: str,
        output_root: str,
        datasource_id: str,
        datasource_root: str,
    ):
        self.datasource_id = datasource_id.strip()
        self.store_file = Path(working_dir).resolve() / "chat_sessions.json"
        self.export_dir = Path(output_root).resolve() / "exports" / "chat"
        self.datasource_root = Path(datasource_root).resolve()
        self._state = self._load_state()

    def get_state(self) -> Dict[str, Any]:
        sessions = [
            {
                **session,
                "messages": [dict(message) for message in session.get("messages", [])],
            }
            for session in self._state["sessions"]
        ]
        return {
            "datasource_id": self.datasource_id,
            "active_session_id": self._state.get("active_session_id"),
            "sessions": sessions,
        }

    def replace_state(
        self, sessions: List[Dict[str, Any]], active_session_id: Optional[str]
    ) -> Dict[str, Any]:
        normalized_sessions = self._normalize_sessions(sessions)
        valid_ids = {session["id"] for session in normalized_sessions}
        next_active_session_id = str(active_session_id).strip() if active_session_id else ""
        if next_active_session_id not in valid_ids:
            next_active_session_id = (
                normalized_sessions[0]["id"] if normalized_sessions else None
            )

        self._state = {
            "datasource_id": self.datasource_id,
            "active_session_id": next_active_session_id,
            "sessions": normalized_sessions,
        }
        self._save()
        return self.get_state()

    def export_markdown(
        self, session_id: str, message_id: Optional[str] = None
    ) -> Dict[str, Any]:
        normalized_session_id = str(session_id).strip()
        session = next(
            (
                item
                for item in self._state["sessions"]
                if item.get("id") == normalized_session_id
            ),
            None,
        )
        if session is None:
            raise KeyError("chat session not found")

        self.export_dir.mkdir(parents=True, exist_ok=True)
        base_name = _sanitize_chat_filename(session.get("title") or "chat-session")
        session_suffix = _sanitize_chat_filename(normalized_session_id)

        if message_id:
            normalized_message_id = str(message_id).strip()
            markdown = _build_single_answer_markdown(
                datasource_id=self.datasource_id,
                session=session,
                message_id=normalized_message_id,
            )
            message_suffix = _sanitize_chat_filename(normalized_message_id)
            file_name = f"{base_name}__answer_{message_suffix}.md"
        else:
            markdown = _build_session_markdown(
                datasource_id=self.datasource_id,
                session=session,
            )
            file_name = f"{base_name}__session_{session_suffix}.md"

        export_path = self.export_dir / file_name
        export_path.write_text(markdown, encoding="utf-8")

        return {
            "status": "ok",
            "datasource_id": self.datasource_id,
            "file_name": file_name,
            "path": str(export_path),
            "relative_path": self._build_export_relative_path(export_path),
            "content": markdown,
        }

    def _load_state(self) -> Dict[str, Any]:
        payload = load_json(str(self.store_file)) or {}
        sessions = payload.get("sessions") if isinstance(payload, dict) else []
        active_session_id = (
            payload.get("active_session_id") if isinstance(payload, dict) else None
        )
        normalized_sessions = self._normalize_sessions(sessions or [])
        valid_ids = {session["id"] for session in normalized_sessions}
        next_active_session_id = (
            str(active_session_id).strip() if active_session_id else None
        )
        if next_active_session_id not in valid_ids:
            next_active_session_id = (
                normalized_sessions[0]["id"] if normalized_sessions else None
            )
        return {
            "datasource_id": self.datasource_id,
            "active_session_id": next_active_session_id,
            "sessions": normalized_sessions,
        }

    def _normalize_sessions(self, sessions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        normalized: List[Dict[str, Any]] = []
        seen_ids: set[str] = set()
        now_ms = int(time.time() * 1000)

        for raw_session in sessions or []:
            if not isinstance(raw_session, dict):
                continue

            session_id = str(raw_session.get("id") or f"session-{uuid4().hex}").strip()
            if not session_id or session_id in seen_ids:
                session_id = f"session-{uuid4().hex}"
            seen_ids.add(session_id)

            messages = self._normalize_messages(raw_session.get("messages"))
            manual_title = bool(raw_session.get("manualTitle"))
            derived_title = _derive_chat_session_title(messages)
            title = str(raw_session.get("title") or "").strip() or "当前对话"
            if not manual_title and derived_title:
                title = derived_title

            updated_at = raw_session.get("updatedAt")
            if not isinstance(updated_at, (int, float)):
                updated_at = now_ms

            normalized.append(
                {
                    "id": session_id,
                    "title": title,
                    "updatedAt": int(updated_at),
                    "manualTitle": manual_title,
                    "messages": messages,
                }
            )

        return normalized

    def _normalize_messages(self, messages: Any) -> List[Dict[str, Any]]:
        normalized: List[Dict[str, Any]] = []
        for raw_message in messages or []:
            if not isinstance(raw_message, dict):
                continue

            role = str(raw_message.get("role") or "").strip()
            if role not in {"user", "assistant"}:
                continue

            payload = {
                key: raw_message[key]
                for key in self._ALLOWED_MESSAGE_KEYS
                if key in raw_message
            }
            payload["id"] = str(payload.get("id") or f"message-{uuid4().hex}")
            payload["role"] = role
            payload["content"] = str(payload.get("content") or "")
            if "timestamp" in payload:
                payload["timestamp"] = str(payload.get("timestamp") or "")
            if "latencyMs" in payload:
                try:
                    payload["latencyMs"] = int(payload["latencyMs"])
                except (TypeError, ValueError):
                    payload.pop("latencyMs", None)
            payload["streaming"] = bool(payload.get("streaming", False))
            payload["error"] = bool(payload.get("error", False))
            normalized.append(payload)
        return normalized

    def _save(self) -> None:
        self.store_file.parent.mkdir(parents=True, exist_ok=True)
        write_json(self._state, str(self.store_file))

    def _build_export_relative_path(self, export_path: Path) -> str:
        repo_like_root = (
            self.datasource_root.parent.parent
            if self.datasource_root.parent.name == "datasources"
            else None
        )

        for base in (repo_like_root, self.datasource_root):
            if base is None:
                continue
            try:
                return export_path.relative_to(base).as_posix()
            except ValueError:
                continue

        try:
            return export_path.relative_to(self.export_dir.parent.parent).as_posix()
        except ValueError:
            return export_path.name


def _derive_chat_session_title(messages: List[Dict[str, Any]]) -> str:
    for message in messages:
        if message.get("role") != "user":
            continue
        content = str(message.get("content") or "").strip()
        if content:
            return content[:16]
    return "当前对话"


def _sanitize_chat_filename(value: str) -> str:
    normalized = re.sub(r"\s+", "-", str(value).strip())
    normalized = re.sub(r'[\\/:*?"<>|]+', "-", normalized)
    normalized = normalized.strip("-._")
    return normalized[:48] or "chat-session"


def _format_export_datetime() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _format_session_updated_at(value: Any) -> str:
    if isinstance(value, (int, float)):
        try:
            return datetime.fromtimestamp(value / 1000).strftime("%Y-%m-%d %H:%M:%S")
        except (ValueError, OSError):
            return "--"
    return "--"


def _truncate_export_evidence_text(value: str, limit: int = 500) -> str:
    normalized = re.sub(r"\s+", " ", str(value or "")).strip()
    if len(normalized) <= limit:
        return normalized
    return normalized[:limit].rstrip() + "..."


def _collect_message_sources(message: Dict[str, Any]) -> List[str]:
    evidence_items = message.get("evidenceItems")
    if isinstance(evidence_items, list):
        sources: List[str] = []
        for item in evidence_items:
            if not isinstance(item, dict):
                continue

            raw = str(item.get("raw") or "").strip()
            snippet = str(item.get("snippet") or "").strip()
            title = str(item.get("title") or "").strip()
            page = str(item.get("page") or "").strip()

            body = _truncate_export_evidence_text(raw or snippet)
            if body:
                label = f"{title} p.{page}" if title and page else title
                if label:
                    sources.append(f"{label}: {body}")
                else:
                    sources.append(body)
                continue

            fallback = str(item.get("sourceId") or title).strip()
            if fallback:
                sources.append(fallback)

        return list(dict.fromkeys(sources))

    context_sources = message.get("contextSources")
    if isinstance(context_sources, list):
        return [str(item) for item in context_sources if str(item).strip()]

    return []


def _append_assistant_sections(lines: List[str], message: Dict[str, Any]) -> None:
    lines.append("### 助手")
    lines.append("")
    lines.append(str(message.get("content") or "").strip() or "（空）")
    lines.append("")

    sources = _collect_message_sources(message)
    if sources:
        lines.append("#### 参考来源")
        lines.append("")
        lines.extend([f"- {source}" for source in sources])
        lines.append("")


def _build_session_markdown(datasource_id: str, session: Dict[str, Any]) -> str:
    lines = [
        "# 历史对话导出",
        "",
        f"- datasource_id: {datasource_id}",
        f"- 会话标题: {str(session.get('title') or '当前对话')}",
        f"- 导出时间: {_format_export_datetime()}",
        "",
    ]

    turn_index = 0
    for message in session.get("messages", []):
        if not isinstance(message, dict):
            continue
        role = message.get("role")
        if role == "user":
            turn_index += 1
            lines.append(f"## 第{turn_index}轮")
            lines.append("")
            lines.append("### 用户")
            lines.append("")
            lines.append(str(message.get("content") or "").strip() or "（空）")
            lines.append("")
        elif role == "assistant":
            if turn_index == 0:
                turn_index = 1
                lines.append(f"## 第{turn_index}轮")
                lines.append("")
            _append_assistant_sections(lines, message)

    return "\n".join(lines).strip() + "\n"


def _build_single_answer_markdown(
    datasource_id: str, session: Dict[str, Any], message_id: str
) -> str:
    messages = session.get("messages", [])
    message_index = next(
        (
            index
            for index, message in enumerate(messages)
            if isinstance(message, dict) and message.get("id") == message_id
        ),
        -1,
    )
    if message_index == -1:
        raise KeyError("chat message not found")

    message = messages[message_index]
    if not isinstance(message, dict) or message.get("role") != "assistant":
        raise ValueError("only assistant messages can be exported")

    previous_user = next(
        (
            candidate
            for candidate in reversed(messages[:message_index])
            if isinstance(candidate, dict) and candidate.get("role") == "user"
        ),
        None,
    )

    lines = [
        "# 单条回答导出",
        "",
        f"- datasource_id: {datasource_id}",
        f"- 会话标题: {str(session.get('title') or '当前对话')}",
        f"- 导出时间: {_format_export_datetime()}",
        "",
    ]

    if isinstance(previous_user, dict):
        lines.append("## 问题")
        lines.append("")
        lines.append(str(previous_user.get("content") or "").strip() or "（空）")
        lines.append("")

    lines.append("## 回答")
    lines.append("")
    _append_assistant_sections(lines, message)
    return "\n".join(lines).strip() + "\n"


def _faultcase_record_text(value: Any) -> str:
    text = str(value or "").strip()
    return text if text else "无"


def build_pipeline_record_source_ids(payload: Dict[str, Any]) -> set[str]:
    records = payload.get("records")
    if not isinstance(records, list):
        return set()

    doc_name = str(payload.get("doc_name") or "").strip()
    if not doc_name:
        return set()

    source_ids: set[str] = set()
    for index, record in enumerate(records):
        if not isinstance(record, dict):
            continue
        record_id = _faultcase_record_text(
            record.get("record_id") or f"llm_{index:03d}"
        )
        source_ids.add(
            compute_mdhash_id(f"{doc_name}::{record_id}", prefix="chunk-faultcase-")
        )
    return source_ids


def _graphml_data_keys(root: ET.Element, target: str) -> Dict[str, str]:
    keys: Dict[str, str] = {}
    for key in root.findall(".//{*}key"):
        if key.attrib.get("for") != target:
            continue
        attr_name = key.attrib.get("attr.name")
        key_id = key.attrib.get("id")
        if attr_name and key_id:
            keys[attr_name] = key_id
    return keys


def _graphml_data_value(element: ET.Element, key_id: Optional[str]) -> str:
    if not key_id:
        return ""
    for data in element.findall("{*}data"):
        if data.attrib.get("key") == key_id:
            return data.text or ""
    return ""


def _split_graph_source_ids(raw_value: str) -> set[str]:
    return {
        item.strip()
        for item in str(raw_value or "").split("<SEP>")
        if item and item.strip()
    }


EXPANDABLE_FAULT_PROPERTY_LABELS = {
    "故障现象",
    "可能原因",
    "处理步骤",
    "注意事项",
    "关键部件",
}

FAULTCARD_SECTION_ALIASES = {
    "记录ID": {"记录ID"},
    "来源路径": {"来源路径", "breadcrumb"},
    "章节标题": {"章节标题"},
    "装备": {"装备"},
    "故障卡片": {"故障卡片"},
    "故障现象": {"故障现象", "现象"},
    "可能原因": {"可能原因", "原因"},
    "处理步骤": {"处理步骤", "处理建议", "维修方法"},
    "可能后果": {"可能后果", "后果"},
    "注意事项": {"注意事项", "注意"},
    "关键部件": {"关键部件"},
    "原始文本": {"原始文本"},
}


def _faultcard_section_for_alias(label: str) -> Optional[str]:
    for section, aliases in FAULTCARD_SECTION_ALIASES.items():
        if label in aliases:
            return section
    return None


def _is_meaningful_property_value(value: Any) -> bool:
    text = str(value or "").strip()
    return bool(text and text not in {"-", "—", "无", "无。", "N/A", "n/a", "null"})


def _structured_field_paragraphs(raw_value: str) -> List[str]:
    paragraphs: List[str] = []
    for line in str(raw_value or "").splitlines():
        text = line.strip()
        if text:
            paragraphs.append(text)
    if paragraphs:
        return paragraphs
    text = str(raw_value or "").strip()
    return [text] if text else []


def _parse_bracketed_faultcard_sections(text: str) -> List[Dict[str, Any]]:
    if "[" not in text or "]" not in text:
        return []

    sections: List[Dict[str, Any]] = []
    pattern = re.compile(r"\[([^\]]+)\]\s*\n?([\s\S]*?)(?=\n\s*\[[^\]]+\]\s*\n?|$)")
    for match in pattern.finditer(text):
        raw_label = str(match.group(1) or "").strip()
        raw_value = str(match.group(2) or "").strip()
        section = _faultcard_section_for_alias(raw_label)
        if not section or not raw_value:
            continue
        sections.append(
            {"label": section, "paragraphs": _structured_field_paragraphs(raw_value)}
        )
    return sections


def _parse_structured_faultcard_sections(text: str) -> List[Dict[str, Any]]:
    normalized_text = re.sub(r"\s+", " ", str(text or "")).strip()
    if not normalized_text:
        return []

    aliases = sorted(
        {alias for aliases in FAULTCARD_SECTION_ALIASES.values() for alias in aliases},
        key=len,
        reverse=True,
    )
    label_pattern = "|".join(re.escape(alias) for alias in aliases)
    pattern = re.compile(
        rf"({label_pattern})\s*[:：]\s*([\s\S]*?)(?=(?:{label_pattern})\s*[:：]|$)"
    )

    sections: List[Dict[str, Any]] = []
    for match in pattern.finditer(normalized_text):
        raw_label = str(match.group(1) or "").strip()
        raw_value = str(match.group(2) or "").strip()
        section = _faultcard_section_for_alias(raw_label)
        if not section or not raw_value:
            continue
        sections.append(
            {"label": section, "paragraphs": _structured_field_paragraphs(raw_value)}
        )
    return sections


def extract_expandable_fault_sections(text: str) -> List[Dict[str, str]]:
    normalized_text = str(text or "").strip()
    if not normalized_text:
        return []

    bracketed_sections = _parse_bracketed_faultcard_sections(normalized_text)
    parsed_sections = (
        bracketed_sections
        if bracketed_sections
        else _parse_structured_faultcard_sections(normalized_text)
    )
    sections: List[Dict[str, str]] = []
    for section in parsed_sections:
        label = str(section.get("label") or "").strip()
        value = "\n".join(section.get("paragraphs") or []).strip()
        if label in EXPANDABLE_FAULT_PROPERTY_LABELS and _is_meaningful_property_value(
            value
        ):
            sections.append({"label": label, "value": value})
    return sections


def build_graphml_source_index(graph_path: Path) -> Optional[Dict[str, Any]]:
    if not graph_path.exists():
        return None
    try:
        root = ET.parse(graph_path).getroot()
    except Exception as graph_exc:
        logging.warning(f"Failed to parse graph stats from {graph_path}: {graph_exc}")
        return None

    node_keys = _graphml_data_keys(root, "node")
    edge_keys = _graphml_data_keys(root, "edge")
    node_source_key = node_keys.get("source_id")
    edge_source_key = edge_keys.get("source_id")
    index: Dict[str, Any] = {"nodes": {}, "edges": {}, "node_attrs": {}}

    for node in root.findall(".//{*}node"):
        node_id = node.attrib.get("id")
        if not node_id:
            continue
        index["node_attrs"][node_id] = {
            "entity_type": _graphml_data_value(node, node_keys.get("entity_type")),
            "description": _graphml_data_value(node, node_keys.get("description")),
            "source_id": _graphml_data_value(node, node_source_key),
        }
        for source_id in _split_graph_source_ids(
            _graphml_data_value(node, node_source_key)
        ):
            index["nodes"].setdefault(source_id, set()).add(node_id)

    for edge_index, edge in enumerate(root.findall(".//{*}edge")):
        edge_id = edge.attrib.get("id") or (
            f"{edge.attrib.get('source', '')}->"
            f"{edge.attrib.get('target', '')}#{edge_index}"
        )
        for source_id in _split_graph_source_ids(
            _graphml_data_value(edge, edge_source_key)
        ):
            index["edges"].setdefault(source_id, set()).add(edge_id)

    return index


def count_graphml_sources(
    graph_source_index: Optional[Dict[str, Any]],
    source_ids: set[str],
) -> tuple[Optional[int], Optional[int]]:
    if graph_source_index is None:
        return None, None

    nodes: set[str] = set()
    edges: set[str] = set()
    for source_id in source_ids:
        nodes.update(graph_source_index["nodes"].get(source_id, set()))
        edges.update(graph_source_index["edges"].get(source_id, set()))

    synthetic_nodes: set[str] = set()
    synthetic_edges: set[str] = set()
    node_attrs = graph_source_index.get("node_attrs", {})
    for node_id in nodes:
        attrs = node_attrs.get(node_id, {})
        entity_type = str(attrs.get("entity_type") or "").replace('"', "").upper()
        if entity_type != "FAULTCASE":
            continue
        description = str(attrs.get("description") or "").strip()
        for section in extract_expandable_fault_sections(description):
            synthetic_id = f"fault-property::{node_id}::{section['label']}"
            synthetic_nodes.add(synthetic_id)
            synthetic_edges.add(f"{node_id}->{synthetic_id}")

    return len(nodes) + len(synthetic_nodes), len(edges) + len(synthetic_edges)


def normalize_document_identity_token(value: Any) -> str:
    text = str(value or "").strip().replace("\\", "/")
    if not text:
        return ""
    name = Path(text).name
    suffix = Path(name).suffix
    if suffix.lower() in {".md", ".txt", ".xlsx", ".json"}:
        name = Path(name).stem
    return re.sub(r"\s+", "", name).lower()


def document_identity_tokens(item: Dict[str, Any]) -> set[str]:
    tokens: set[str] = set()
    for key in ("name", "doc_id", "relative_path", "absolute_path"):
        token = normalize_document_identity_token(item.get(key))
        if token:
            tokens.add(token)
    return tokens


def find_matching_pipeline_item(
    source_item: Dict[str, Any], pipeline_items: List[Dict[str, Any]]
) -> Optional[Dict[str, Any]]:
    source_tokens = document_identity_tokens(source_item)
    if not source_tokens:
        return None
    for pipeline_item in pipeline_items:
        if source_tokens & document_identity_tokens(pipeline_item):
            return pipeline_item
    return None


def merge_source_item_with_pipeline_graph_counts(
    source_item: Dict[str, Any], pipeline_item: Optional[Dict[str, Any]]
) -> Dict[str, Any]:
    if pipeline_item is None:
        enriched = dict(source_item)
        enriched.setdefault("ready_to_query", False)
        return enriched

    enriched = dict(source_item)
    enriched["ready_to_query"] = True
    enriched["status"] = "indexed"
    enriched["error"] = None
    if pipeline_item.get("indexed_at") is not None:
        enriched["indexed_at"] = pipeline_item.get("indexed_at")
    for key in ("chunks_count", "graph_nodes", "graph_edges"):
        if pipeline_item.get(key) is not None:
            enriched[key] = pipeline_item.get(key)
    enriched["pipeline_record_count"] = pipeline_item.get("chunks_count")
    enriched["pipeline_relative_path"] = pipeline_item.get("relative_path")
    enriched["pipeline_modified_at"] = pipeline_item.get("modified_at")
    return enriched


def build_pipeline_outputs_summary(datasource: ResolvedDatasource) -> Dict[str, Any]:
    items: List[Dict[str, Any]] = []
    tombstones = _load_pipeline_tombstones(datasource)
    records_root = datasource.staging_root / "extracted" / "records"
    if not records_root.exists():
        return {
            "datasource_id": datasource.id,
            "stats": {
                "total": 0,
                "indexed": 0,
                "processing": 0,
                "failed": 0,
                "pending": 0,
            },
            "items": [],
        }

    graph_source_index = build_graphml_source_index(
        datasource.working_dir / "graph_chunk_entity_relation.graphml"
    )

    for records_path in sorted(records_root.rglob("diagnostic_records_llm.json")):
        if not records_path.exists():
            continue

        try:
            payload = load_json(str(records_path)) or {}
            stat = records_path.stat()
            payload_doc_name = str(payload.get("doc_name") or records_path.parent.name)
            if records_path.parent.name in tombstones or payload_doc_name in tombstones:
                continue
            graph_nodes, graph_edges = count_graphml_sources(
                graph_source_index,
                build_pipeline_record_source_ids(payload),
            )

            items.append(
                {
                    "datasource_id": datasource.id,
                    "name": payload_doc_name,
                    "relative_path": str(records_path.relative_to(REPO_ROOT)),
                    "absolute_path": str(records_path.resolve()),
                    "type": "RECORDS",
                    "size": stat.st_size,
                    "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "status": "indexed",
                    "doc_id": payload_doc_name,
                    "chunks_count": payload.get("record_count", 0),
                    "graph_nodes": graph_nodes,
                    "graph_edges": graph_edges,
                    "content_length": stat.st_size,
                    "content_summary": f"records 文档，包含 {payload.get('record_count', 0)} 条故障卡片",
                    "indexed_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    "source_kind": "pipeline",
                    **build_resource_descriptor("pipeline", datasource.id),
                }
            )
        except Exception as exc:  # noqa: BLE001
            logging.error(f"Failed to summarize pipeline doc {records_path}: {exc}")

    stats = {
        "total": len(items),
        "indexed": len(items),
        "processing": 0,
        "failed": 0,
        "pending": 0,
    }
    return {"datasource_id": datasource.id, "stats": stats, "items": items}


# Pydantic models
class SearchMode(str, Enum):
    graph_text_hybrid = "graph_text_hybrid"
    graph_only = "graph_only"
    text_only = "text_only"
    keyword_search = "keyword_search"
    faultcase_fast = "faultcase_fast"


class OllamaMessage(BaseModel):
    role: str
    content: str
    images: Optional[List[str]] = None


class OllamaChatRequest(BaseModel):
    model: str = ollama_server_infos.LIGHTRAG_MODEL
    messages: List[OllamaMessage]
    stream: bool = True  # Default to streaming mode
    options: Optional[Dict[str, Any]] = None
    system: Optional[str] = None


class OllamaChatResponse(BaseModel):
    model: str
    created_at: str
    message: OllamaMessage
    done: bool


class OllamaGenerateRequest(BaseModel):
    model: str = ollama_server_infos.LIGHTRAG_MODEL
    prompt: str
    system: Optional[str] = None
    stream: bool = False
    options: Optional[Dict[str, Any]] = None


class OllamaGenerateResponse(BaseModel):
    model: str
    created_at: str
    response: str
    done: bool
    context: Optional[List[int]]
    total_duration: Optional[int]
    load_duration: Optional[int]
    prompt_eval_count: Optional[int]
    prompt_eval_duration: Optional[int]
    eval_count: Optional[int]
    eval_duration: Optional[int]


class OllamaVersionResponse(BaseModel):
    version: str


class OllamaModelDetails(BaseModel):
    parent_model: str
    format: str
    family: str
    families: List[str]
    parameter_size: str
    quantization_level: str


class OllamaModel(BaseModel):
    name: str
    model: str
    size: int
    digest: str
    modified_at: str
    details: OllamaModelDetails


class OllamaTagResponse(BaseModel):
    models: List[OllamaModel]


class ConversationHistoryMessage(BaseModel):
    role: str
    content: str


class QueryRequest(BaseModel):
    query: str
    mode: SearchMode = SearchMode.graph_text_hybrid
    query_semantics: str = "natural_language"
    keywords: List[str] = Field(default_factory=list)
    text_only_retrieval: bool = False
    stream: bool = False
    only_need_context: bool = False
    conversation_history: List[ConversationHistoryMessage] = Field(
        default_factory=list
    )


class QueryResponse(BaseModel):
    response: str


class ChatSessionStateRequest(BaseModel):
    datasource_id: Optional[str] = None
    active_session_id: Optional[str] = None
    sessions: List[Dict[str, Any]] = Field(default_factory=list)


class ChatSessionStateResponse(BaseModel):
    datasource_id: str
    active_session_id: Optional[str] = None
    sessions: List[Dict[str, Any]] = Field(default_factory=list)


class ChatExportRequest(BaseModel):
    datasource_id: Optional[str] = None
    session_id: str
    message_id: Optional[str] = None


class ChatExportResponse(BaseModel):
    status: str
    datasource_id: str
    file_name: str
    path: str
    relative_path: str
    content: str


class EmptyQueryResponseError(RuntimeError):
    """Raised when the upstream provider returns an empty answer body."""


class AliasRecordResponse(BaseModel):
    datasource_id: str
    id: str
    canonical_name: str
    entity_type: str
    alias: str
    alias_norm: str
    enabled: bool
    reviewed: bool
    created_at: str
    updated_at: str


class AliasListResponse(BaseModel):
    datasource_id: str
    items: List[AliasRecordResponse]
    total: int
    stats: Dict[str, Any]


class AliasCreateRequest(BaseModel):
    datasource_id: Optional[str] = None
    canonical_name: str
    entity_type: str
    alias: str
    enabled: bool = True
    reviewed: bool = False


class AliasUpdateRequest(BaseModel):
    datasource_id: Optional[str] = None
    canonical_name: Optional[str] = None
    entity_type: Optional[str] = None
    alias: Optional[str] = None
    enabled: Optional[bool] = None
    reviewed: Optional[bool] = None


class AliasResolveResponse(BaseModel):
    datasource_id: str
    query: str
    query_norm: str
    intent: str
    preferred_entity_types: List[str]
    alias_hits: List[AliasRecordResponse]


class InsertTextRequest(BaseModel):
    text: str
    description: Optional[str] = None


class InsertResponse(BaseModel):
    status: str
    message: str
    document_count: int


def get_api_key_dependency(api_key: Optional[str]):
    if not api_key:
        # If no API key is configured, return a dummy dependency that always succeeds
        async def no_auth():
            return None

        return no_auth

    # If API key is configured, use proper authentication
    api_key_header = APIKeyHeader(name="X-API-Key", auto_error=False)

    async def api_key_auth(api_key_header_value: str | None = Security(api_key_header)):
        if not api_key_header_value:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="API Key required"
            )
        if api_key_header_value != api_key:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Invalid API Key"
            )
        return api_key_header_value

    return api_key_auth


def create_app(args):
    supported_llm_bindings = [
        "lollms",
        "ollama",
        "openai",
        "deepseek",
        "openai-ollama",
        "azure_openai",
    ]
    supported_embedding_bindings = ["lollms", "ollama", "openai", "azure_openai"]

    if args.llm_binding not in supported_llm_bindings:
        raise Exception("llm binding not supported")

    if args.embedding_binding not in supported_embedding_bindings:
        raise Exception("embedding binding not supported")

    if args.query_llm_binding not in supported_llm_bindings:
        raise Exception("llm binding not supported")
    if args.graph_storage not in ["NetworkXStorage", "Neo4JStorage"]:
        raise Exception("graph storage not supported")

    resolved_datasource = _resolve_args_datasource(args)
    ship_benchmark_adapter.configure_source(
        source_csv_path=resolve_ship_benchmark_source(
            datasource_id=resolved_datasource.id,
            datasource_root=resolved_datasource.root,
            repo_root=REPO_ROOT,
        )
    )

    # Set default hosts if not provided
    if args.llm_binding_host is None:
        args.llm_binding_host = get_default_host(args.llm_binding)

    if args.embedding_binding_host is None:
        args.embedding_binding_host = get_default_host(args.embedding_binding)

    if args.query_llm_binding_host is None:
        args.query_llm_binding_host = get_default_host(args.query_llm_binding)
    if args.graph_storage == "Neo4JStorage":
        if not all([args.neo4j_uri, args.neo4j_username, args.neo4j_password]):
            raise Exception(
                "NEO4J_URI, NEO4J_USERNAME and NEO4J_PASSWORD are required when graph-storage is Neo4JStorage"
            )
        os.environ["NEO4J_URI"] = args.neo4j_uri
        os.environ["NEO4J_USERNAME"] = args.neo4j_username
        os.environ["NEO4J_PASSWORD"] = args.neo4j_password
        os.environ["NEO4J_DATABASE"] = args.neo4j_database
        try:
            asyncio.run(_verify_neo4j_connection(args))
        except Exception as exc:
            if args.graph_storage_fallback == "NetworkXStorage":
                logging.warning(
                    "Neo4j is configured but unavailable at %s; falling back to NetworkXStorage: %s",
                    args.neo4j_uri,
                    exc,
                )
                args.graph_storage = "NetworkXStorage"
            else:
                raise

    # Add SSL validation
    if args.ssl:
        if not args.ssl_certfile or not args.ssl_keyfile:
            raise Exception(
                "SSL certificate and key files must be provided when SSL is enabled"
            )
        if not os.path.exists(args.ssl_certfile):
            raise Exception(f"SSL certificate file not found: {args.ssl_certfile}")
        if not os.path.exists(args.ssl_keyfile):
            raise Exception(f"SSL key file not found: {args.ssl_keyfile}")

    # Setup logging
    logging.basicConfig(
        format="%(levelname)s:%(message)s", level=getattr(logging, args.log_level)
    )

    # Check if API key is provided either through env var or args
    api_key = os.getenv("LIGHTRAG_API_KEY") or args.key
    ollama_server_infos.GRAPH_STORAGE = args.graph_storage

    # Initialize document manager
    doc_manager = DocumentManager(args.input_dir)
    doc_registry = DocumentRegistry(
        args.working_dir, args.input_dir, args.datasource_id
    )
    doc_registry.sync_from_disk(doc_manager.supported_extensions)
    doc_manager.indexed_files = doc_registry.get_indexed_paths()
    active_document_pipelines: Dict[str, Dict[str, Any]] = {}
    active_document_pipelines_lock = threading.Lock()

    def document_pipeline_key(
        datasource: ResolvedDatasource, file_path: Union[str, Path]
    ) -> str:
        return f"{datasource.id}:{Path(file_path).resolve()}"

    def is_pipeline_cancelled(stop_event: threading.Event) -> bool:
        return stop_event.is_set()

    def raise_if_pipeline_cancelled(stop_event: threading.Event) -> None:
        if is_pipeline_cancelled(stop_event):
            raise RuntimeError("document processing cancelled")

    def register_document_pipeline(
        datasource: ResolvedDatasource,
        file_path: Union[str, Path],
        stop_event: threading.Event,
    ) -> str:
        key = document_pipeline_key(datasource, file_path)
        current_task = asyncio.current_task()
        with active_document_pipelines_lock:
            existing = active_document_pipelines.get(key)
            existing_task = existing.get("task") if existing else None
            if existing_task is not None and not existing_task.done():
                raise HTTPException(
                    status_code=409,
                    detail="Document processing is already running",
                )
            active_document_pipelines[key] = {
                "task": current_task,
                "stop_event": stop_event,
                "started_at": datetime.now().isoformat(),
            }
        return key

    def unregister_document_pipeline(key: str, stop_event: threading.Event) -> None:
        with active_document_pipelines_lock:
            current = active_document_pipelines.get(key)
            if current and current.get("stop_event") is stop_event:
                active_document_pipelines.pop(key, None)

    def request_document_pipeline_stop(
        datasource: ResolvedDatasource, file_path: Union[str, Path]
    ) -> bool:
        key = document_pipeline_key(datasource, file_path)
        with active_document_pipelines_lock:
            current = active_document_pipelines.get(key)
            if not current:
                return False
            stop_event = current.get("stop_event")
            if isinstance(stop_event, threading.Event):
                stop_event.set()
            return True

    async def warmup_query_runtime() -> None:
        if os.getenv("SKIP_BACKEND_WARMUP", "false").lower() in {"1", "true", "yes", "on"}:
            logging.info("Backend query/embedding warmup skipped by SKIP_BACKEND_WARMUP")
            return
        warmup_timeout = min(args.timeout or 30, 30)

        try:
            await asyncio.wait_for(
                query_rag.llm_model_func(
                    "请回复：已就绪",
                    system_prompt="你是启动预热探针，只回复“已就绪”。",
                ),
                timeout=warmup_timeout,
            )
            active_models = get_active_model_configs()
            logging.info(
                "Query LLM warmup completed for model=%s host=%s",
                active_models["query"].model,
                active_models["query"].host,
            )
        except Exception as exc:
            logging.warning("Query LLM warmup failed: %s", exc)

        try:
            await asyncio.wait_for(
                query_rag.embedding_func(["启动预热"]),
                timeout=warmup_timeout,
            )
            active_models = get_active_model_configs()
            logging.info(
                "Embedding warmup completed for model=%s host=%s",
                active_models["embedding"].model,
                active_models["embedding"].host,
            )
        except Exception as exc:
            logging.warning("Embedding warmup failed: %s", exc)

    @asynccontextmanager
    async def lifespan(app: FastAPI):
        """Lifespan context manager for startup and shutdown events"""
        # Startup logic
        doc_registry.sync_from_disk(doc_manager.supported_extensions)
        doc_manager.indexed_files = doc_registry.get_indexed_paths()
        warmup_task = asyncio.create_task(warmup_query_runtime())
        if args.auto_scan_at_startup:
            try:
                new_files = doc_manager.scan_directory_for_new_files()
                for i in range(0, len(new_files), args.batch_insert_size):
                    batch_files = new_files[i : i + args.batch_insert_size]
                    try:
                        await index_files(batch_files)
                    except Exception as e:
                        trace_exception(e)
                        logging.error(
                            f"Error indexing files {[str(path) for path in batch_files]}: {str(e)}"
                        )

                ASCIIColors.info(
                    f"Indexed {len(new_files)} documents from {args.input_dir}"
                )
            except Exception as e:
                logging.error(f"Error during startup indexing: {str(e)}")
        yield
        # Cleanup logic (if needed)
        if not warmup_task.done():
            warmup_task.cancel()
            with contextlib.suppress(asyncio.CancelledError):
                await warmup_task

    # Initialize FastAPI
    app = FastAPI(
        title="MiniRAG API",
        description="API for querying text using MiniRAG with separate storage and input directories"
        + "(With authentication)"
        if api_key
        else "",
        version=__api_version__,
        openapi_tags=[{"name": "api"}],
        lifespan=lifespan,
    )

    # Add CORS middleware
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=True,
        allow_methods=["*"],
        allow_headers=["*"],
    )

    # Create the optional API key dependency
    optional_api_key = get_api_key_dependency(api_key)

    # Create working directory if it doesn't exist
    Path(args.working_dir).mkdir(parents=True, exist_ok=True)
    model_registry = load_model_registry(args, args.model_registry)
    active_model_selection = resolve_runtime_selection(model_registry)

    def build_llm_components(binding, binding_host, binding_api_key, model_name):
        if binding == "lollms":
            from minirag.llm.lollms import lollms_model_complete

            return lollms_model_complete, {
                "host": binding_host,
                "timeout": args.timeout,
                "api_key": binding_api_key,
            }
        if binding in ["openai", "deepseek"]:
            from minirag.llm.openai import openai_complete_if_cache

            return openai_complete_if_cache, {
                "model": model_name,
                "base_url": binding_host,
                "timeout": args.timeout,
                "api_key": binding_api_key,
            }
        if binding == "azure_openai":
            from minirag.llm.azure_openai import azure_openai_complete_if_cache

            return azure_openai_complete_if_cache, {
                "base_url": binding_host,
                "timeout": args.timeout,
                "api_key": binding_api_key,
            }
        from minirag.llm.ollama import ollama_model_complete

        return ollama_model_complete, {
            "host": binding_host,
            "timeout": args.timeout,
            "options": {"num_ctx": args.max_tokens},
            "api_key": binding_api_key,
        }

    def build_embedding_func(embedding_model_config):
        if embedding_model_config.binding == "lollms":
            from minirag.llm.lollms import lollms_embed

            embed_callable = lambda texts: lollms_embed(
                texts,
                embed_model=embedding_model_config.model,
                host=embedding_model_config.host,
                api_key=embedding_model_config.api_key,
            )
        elif embedding_model_config.binding == "ollama":
            from minirag.llm.ollama import ollama_embed

            embed_callable = lambda texts: ollama_embed(
                texts,
                embed_model=embedding_model_config.model,
                host=embedding_model_config.host,
                api_key=embedding_model_config.api_key,
            )
        elif embedding_model_config.binding == "azure_openai":
            from minirag.llm.azure_openai import azure_openai_embed

            embed_callable = lambda texts: azure_openai_embed(
                texts,
                model=embedding_model_config.model,
                api_key=embedding_model_config.api_key,
            )
        else:
            from minirag.llm.openai import openai_embed

            embed_callable = lambda texts: openai_embed(
                texts,
                model=embedding_model_config.model,
                base_url=embedding_model_config.host,
                api_key=embedding_model_config.api_key,
            )

        return EmbeddingFunc(
            embedding_dim=args.embedding_dim,
            max_token_size=args.max_embed_tokens,
            func=embed_callable,
        )

    def build_rag(llm_model_config, embedding_model_config):
        llm_model_func, llm_model_kwargs = build_llm_components(
            llm_model_config.binding,
            llm_model_config.host,
            llm_model_config.api_key,
            llm_model_config.model,
        )
        return MiniRAG(
            working_dir=args.working_dir,
            datasource_id=args.datasource_id,
            llm_model_func=llm_model_func,
            llm_model_name=llm_model_config.model,
            llm_model_max_async=args.max_async,
            llm_model_max_token_size=args.max_tokens,
            chunk_token_size=int(args.chunk_size),
            chunk_overlap_token_size=int(args.chunk_overlap_size),
            llm_model_kwargs=llm_model_kwargs,
            embedding_func=build_embedding_func(embedding_model_config),
            kv_storage=ollama_server_infos.KV_STORAGE,
            graph_storage=ollama_server_infos.GRAPH_STORAGE,
            vector_storage=ollama_server_infos.VECTOR_STORAGE,
            doc_status_storage=ollama_server_infos.DOC_STATUS_STORAGE,
            vector_db_storage_cls_kwargs={
                "cosine_better_than_threshold": args.cosine_threshold
            },
            max_parallel_insert=args.max_parallel_insert,
        )

    alias_store = FaultcaseAliasStore(
        args.working_dir, datasource_id=args.datasource_id
    )

    def rebuild_runtime_rags(next_selection: RuntimeModelSelection) -> None:
        nonlocal index_rag, query_rag, active_model_selection

        index_model_config = model_registry.resolve_llm(
            next_selection.index_llm_id, role=LLM_ROLE_INDEX
        )
        query_model_config = model_registry.resolve_llm(
            next_selection.query_llm_id, role=LLM_ROLE_QUERY
        )
        embedding_model_config = model_registry.resolve_embedding(
            next_selection.embedding_model_id
        )

        index_rag = build_rag(index_model_config, embedding_model_config)
        query_rag = build_rag(query_model_config, embedding_model_config)
        query_rag.faultcase_alias_store = alias_store
        active_model_selection = next_selection
        BENCHMARK_AVAILABLE_MODELS[:] = [
            BenchmarkModelOption(
                id=next_selection.query_llm_id,
                label=query_model_config.label,
            )
        ]
        benchmark_state.set_model(next_selection.query_llm_id)

    index_rag = None
    query_rag = None
    rebuild_runtime_rags(active_model_selection)
    query_rag.faultcase_alias_store = alias_store

    chat_session_store = ChatSessionStore(
        args.working_dir,
        str(resolved_datasource.output_root),
        args.datasource_id,
        str(resolved_datasource.root),
    )

    graph_datasource_scope = build_datasource_scope(resolved_datasource)
    equipment_catalog = build_faultcase_equipment_catalog(
        resolved_datasource.output_root
    )

    def get_active_model_catalog() -> RuntimeModelCatalogResponse:
        return build_runtime_model_catalog(model_registry, active_model_selection)

    def get_active_model_configs():
        return {
            "index": model_registry.resolve_llm(
                active_model_selection.index_llm_id, role=LLM_ROLE_INDEX
            ),
            "query": model_registry.resolve_llm(
                active_model_selection.query_llm_id, role=LLM_ROLE_QUERY
            ),
            "embedding": model_registry.resolve_embedding(
                active_model_selection.embedding_model_id
            ),
        }

    def normalize_openai_chat_url(base_url: str) -> str:
        cleaned = str(base_url or "").strip().rstrip("/")
        if cleaned.endswith("/chat/completions"):
            return cleaned
        if cleaned.endswith("/v1"):
            return f"{cleaned}/chat/completions"
        return f"{cleaned}/v1/chat/completions"

    def require_openai_compatible_llm_config():
        query_model_config = model_registry.resolve_llm(
            active_model_selection.query_llm_id, role=LLM_ROLE_QUERY
        )
        if query_model_config.binding not in {"openai", "deepseek", "azure_openai"}:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Ready-to-query pipeline requires an OpenAI-compatible local "
                    f"query LLM endpoint; current binding is {query_model_config.binding!r}."
                ),
            )
        if not query_model_config.host:
            raise HTTPException(
                status_code=400,
                detail="Ready-to-query pipeline requires a query LLM host/base_url.",
            )
        return query_model_config

    def build_graph_pipeline_args(request_datasource: ResolvedDatasource, doc_dir_name: str):
        query_model_config = require_openai_compatible_llm_config()
        embedding_model_config = model_registry.resolve_embedding(
            active_model_selection.embedding_model_id
        )
        embedding_binding = (
            "ollama" if embedding_model_config.binding == "ollama" else "openai"
        )
        return argparse.Namespace(
            datasource_id=request_datasource.id,
            datasource_root=str(request_datasource.root),
            all_docs=False,
            doc_dirs=[doc_dir_name],
            exclude_doc=[],
            working_dir=str(request_datasource.working_dir),
            embedding_binding=embedding_binding,
            embedding_model=embedding_model_config.model,
            embedding_host=embedding_model_config.host or "",
            embedding_api_key=embedding_model_config.api_key or "lm-studio",
            embedding_dim=args.embedding_dim,
            parent_node_mode=os.getenv("READY_PIPELINE_PARENT_NODE_MODE", "hybrid"),
            parent_scope_mode=os.getenv("READY_PIPELINE_PARENT_SCOPE_MODE", "auto"),
            parent_llm_url=normalize_openai_chat_url(query_model_config.host),
            parent_llm_model=query_model_config.model,
            parent_llm_api_key=query_model_config.api_key or "lm-studio",
            parent_llm_timeout=int(os.getenv("READY_PIPELINE_PARENT_LLM_TIMEOUT", "120")),
            log_level=os.getenv("READY_PIPELINE_GRAPH_LOG_LEVEL", "INFO"),
            reset_working_dir=False,
        )

    async def run_ready_to_query_pipeline(
        *,
        request_datasource: ResolvedDatasource,
        file_path: Path,
        stop_event: threading.Event,
    ) -> dict[str, Any]:
        doc_dir_name = derive_doc_dir_name(file_path)
        doc_dir = request_datasource.staging_root / "chunks" / doc_dir_name
        good_chunks_dir = doc_dir / "good_chunks"
        raise_if_pipeline_cancelled(stop_event)
        if not good_chunks_dir.exists() or not any(good_chunks_dir.glob("*.json")):
            raise HTTPException(
                status_code=409,
                detail=(
                    "No routed good_chunks found for this document. Upload/reindex it "
                    "or run the existing chunk-only reprocess action first."
                ),
            )

        query_model_config = require_openai_compatible_llm_config()
        extract_module = importlib.import_module(
            "data_pipeline.scripts.03_extract_fault_records_llm"
        )
        aggregate_module = importlib.import_module(
            "data_pipeline.scripts.04_aggregate_llm_records"
        )
        cards_module = importlib.import_module(
            "data_pipeline.scripts.05_export_faultcase_txt_cards"
        )
        graph_module = importlib.import_module(
            "data_pipeline.scripts.06_build_faultcase_graph"
        )

        extract_summary = await asyncio.to_thread(
            extract_module.screen_one_doc,
            doc_dir=doc_dir,
            input_folder="good_chunks",
            accepted_folder="accepted_records",
            rejected_folder="rejected_chunks",
            report_name="llm_extract_report.json",
            api_key=query_model_config.api_key or "lm-studio",
            base_url=query_model_config.host,
            model=query_model_config.model,
            temperature=float(os.getenv("READY_PIPELINE_EXTRACT_TEMPERATURE", "0")),
            max_retries=int(os.getenv("READY_PIPELINE_EXTRACT_MAX_RETRIES", "3")),
            retry_delay=float(os.getenv("READY_PIPELINE_EXTRACT_RETRY_DELAY", "2")),
            max_concurrency=int(
                os.getenv("READY_PIPELINE_EXTRACT_MAX_CONCURRENCY", "1")
            ),
            requests_per_minute=int(
                os.getenv("READY_PIPELINE_EXTRACT_REQUESTS_PER_MINUTE", "120")
            ),
            write_report_file=True,
            verbose=os.getenv("READY_PIPELINE_VERBOSE", "").lower()
            in {"1", "true", "yes"},
            stop_event=stop_event,
        )
        raise_if_pipeline_cancelled(stop_event)

        aggregate_args = argparse.Namespace(
            datasource_id=request_datasource.id,
            datasource_root=str(request_datasource.root),
            doc_dir=doc_dir_name,
            accepted_folder="accepted_records",
            good_folder="good_chunks",
            write_review=True,
            cleanup_intermediates=False,
        )
        aggregate_summary = await asyncio.to_thread(
            aggregate_module.aggregate_one_doc, aggregate_args
        )
        raise_if_pipeline_cancelled(stop_event)

        records_path = Path(aggregate_summary["records_path"])
        cards_output_dir = (
            request_datasource.output_root / "exports" / "cards" / doc_dir_name
        )
        cards_summary = await asyncio.to_thread(
            cards_module.export_faultcase_txt_cards,
            records_path,
            cards_output_dir,
        )
        raise_if_pipeline_cancelled(stop_event)

        await graph_module.build_graph(
            build_graph_pipeline_args(request_datasource, doc_dir_name)
        )
        raise_if_pipeline_cancelled(stop_event)
        rebuild_runtime_rags(active_model_selection)

        return {
            "doc_dir": str(doc_dir),
            "records_path": str(records_path),
            "cards_dir": cards_summary["cards_dir"],
            "accepted_count": extract_summary.get("accepted_chunks", 0),
            "rejected_count": extract_summary.get("rejected_chunks", 0),
            "kept_record_count": aggregate_summary.get("kept_record_count", 0),
        }

    def resolve_graph_datasource_request(
        datasource_id: Optional[str] = None,
        datasource_root: Optional[str] = None,
        input_dir: Optional[str] = None,
        working_dir: Optional[str] = None,
        outputs_root: Optional[str] = None,
    ) -> ResolvedDatasource:
        selector_kwargs = {
            "datasource_id": _normalize_datasource_selector(datasource_id),
            "datasource_root": _normalize_datasource_selector(datasource_root),
            "input_dir": _normalize_datasource_selector(input_dir),
            "working_dir": _normalize_datasource_selector(working_dir),
            "outputs_root": _normalize_datasource_selector(outputs_root),
        }
        if not any(selector_kwargs.values()):
            raise HTTPException(
                status_code=400,
                detail="A datasource selector is required; provide datasource_id, datasource_root, or a legacy path selector.",
            )

        try:
            requested = resolve_datasource(repo_root=REPO_ROOT, **selector_kwargs)
        except DatasourceResolutionError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

        if requested.id != resolved_datasource.id:
            raise HTTPException(status_code=404, detail="Datasource not found")

        return requested

    def require_graph_datasource(
        datasource_id: Optional[str] = None,
        datasource_root: Optional[str] = None,
        input_dir: Optional[str] = None,
        working_dir: Optional[str] = None,
        outputs_root: Optional[str] = None,
    ) -> str:
        return resolve_graph_datasource_request(
            datasource_id=datasource_id,
            datasource_root=datasource_root,
            input_dir=input_dir,
            working_dir=working_dir,
            outputs_root=outputs_root,
        ).id

    def require_alias_datasource(datasource_id: Optional[str] = None) -> str:
        requested = str(datasource_id or args.datasource_id or "").strip()
        if not requested:
            raise HTTPException(status_code=400, detail="datasource_id is required")
        if requested != args.datasource_id:
            raise HTTPException(status_code=404, detail="Datasource not found")
        return requested

    def build_doc_id(content: str) -> str:
        return compute_mdhash_id(clean_text(content), prefix="doc-")

    def build_query_param(
        request: QueryRequest, *, provider_stream: bool
    ) -> QueryParam:
        shared_profile = build_shared_query_profile()
        return QueryParam(
            mode=request.mode,
            query_semantics=(
                "keyword_search"
                if request.query_semantics == "keyword_search"
                or request.mode == SearchMode.keyword_search
                else "natural_language"
            ),
            keywords=[str(item).strip() for item in request.keywords if str(item).strip()],
            text_only_retrieval=bool(request.text_only_retrieval),
            stream=provider_stream,
            only_need_context=request.only_need_context,
            # Temporarily ignore frontend-provided history to restore
            # backend single-turn behavior without changing the API contract.
            # conversation_history=[
            #     {"role": item.role, "content": item.content}
            #     for item in request.conversation_history
            # ],
            conversation_history=[],
            # history_turns=args.history_turns,
            history_turns=0,
            top_k=shared_profile["top_k"],
            response_type=shared_profile["response_type"],
            max_token_for_text_unit=shared_profile["max_token_for_text_unit"],
            max_token_for_node_context=shared_profile["max_token_for_node_context"],
        )

    def build_shared_query_profile() -> dict[str, Any]:
        return {
            "top_k": min(args.top_k, int(os.getenv("BENCHMARK_TOP_K", "8"))),
            "max_token_for_text_unit": int(
                os.getenv("BENCHMARK_MAX_TOKEN_FOR_TEXT_UNIT", "8000")
            ),
            "max_token_for_node_context": int(
                os.getenv("BENCHMARK_MAX_TOKEN_FOR_NODE_CONTEXT", "300")
            ),
            "response_type": os.getenv("BENCHMARK_RESPONSE_TYPE", "Bullet List"),
        }

    async def execute_query_request(
        request: QueryRequest, *, provider_stream: bool
    ) -> Any:
        return await query_rag.aquery(
            request.query,
            param=build_query_param(request, provider_stream=provider_stream),
        )

    def ensure_non_empty_query_text(text: Any) -> str:
        normalized = str(text or "")
        if normalized.strip():
            return normalized
        raise EmptyQueryResponseError(
            "Upstream model returned an empty response body"
        )

    async def collect_query_response_text(response: Any) -> str:
        if isinstance(response, str):
            return ensure_non_empty_query_text(response)

        chunks: list[str] = []
        async for chunk in response:
            if chunk:
                chunks.append(str(chunk))

        return ensure_non_empty_query_text("".join(chunks))

    def build_plain_streaming_response(content_generator):
        from fastapi.responses import StreamingResponse

        return StreamingResponse(
            content_generator,
            media_type="text/plain; charset=utf-8",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "POST, OPTIONS",
                "Access-Control-Allow-Headers": "Content-Type",
                "X-Accel-Buffering": "no",
            },
        )

    async def get_doc_status_snapshot(doc_id: str) -> Optional[Dict[str, Any]]:
        try:
            payload = await index_rag.doc_status.get_by_id(doc_id)
            if isinstance(payload, dict):
                return payload
            if payload is not None and hasattr(payload, "__dict__"):
                return dict(payload.__dict__)
            return None
        except Exception:
            return None

    def merge_inventory_item_with_doc_status(
        item: Dict[str, Any], doc_status: Optional[Dict[str, Any]]
    ) -> Dict[str, Any]:
        if not doc_status:
            return item

        merged = dict(item)
        runtime_status = str(doc_status.get("status") or "").strip().lower()

        if runtime_status == "processed":
            merged["status"] = "indexed"
            merged["error"] = None
        elif runtime_status == "processing":
            merged["status"] = "processing"
            merged["error"] = None
        elif runtime_status == "pending":
            merged["status"] = "pending"
        elif runtime_status == "failed":
            merged["status"] = "failed"
            merged["error"] = doc_status.get("error") or merged.get("error")

        for key in (
            "chunks_count",
            "content_length",
            "content_summary",
            "created_at",
            "updated_at",
        ):
            value = doc_status.get(key)
            if value is not None:
                merged[key] = value

        merged["doc_status"] = runtime_status or merged.get("doc_status")
        return merged

    async def enrich_inventory_item_fallback_summary(
        item: Dict[str, Any]
    ) -> Dict[str, Any]:
        if item.get("source_kind") == "pipeline":
            return item
        if item.get("content_summary"):
            return item

        absolute_path = item.get("absolute_path")
        if not absolute_path:
            return item

        file_path = Path(absolute_path)
        if not file_path.exists():
            return item

        try:
            content = await load_file_content(file_path)
        except Exception:
            return item

        if not content:
            return item

        enriched = dict(item)
        enriched["content_summary"] = get_content_summary(content)
        enriched["content_length"] = len(content)
        if not enriched.get("status"):
            enriched["status"] = "indexed"
            enriched["error"] = None
            enriched["indexed_at"] = enriched.get("indexed_at") or enriched.get(
                "modified_at"
            )
        return enriched
        return enriched

    async def load_file_content(file_path: Union[str, Path]) -> str:
        """Load supported file types into plain text content."""
        if not pm.is_installed("aiofiles"):
            pm.install("aiofiles")

        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        content = ""
        ext = file_path.suffix.lower()

        match ext:
            case ".txt" | ".md":
                async with aiofiles.open(file_path, "r", encoding="utf-8") as f:
                    content = await f.read()

            case ".xlsx":
                if not pm.is_installed("openpyxl"):
                    pm.install("openpyxl")
                from openpyxl import load_workbook

                workbook = load_workbook(filename=str(file_path), data_only=True)
                sheet_blocks: list[str] = []
                for sheet in workbook.worksheets:
                    rows: list[str] = []
                    for row in sheet.iter_rows(values_only=True):
                        cells = [str(value).strip() for value in row if value is not None and str(value).strip()]
                        if cells:
                            rows.append(" | ".join(cells))
                    if rows:
                        sheet_blocks.append(f"# {sheet.title}\n" + "\n".join(rows))
                content = "\n\n".join(sheet_blocks)

            case _:
                raise ValueError(f"Unsupported file format: {ext}")

        return content

    def sync_source_document_pipeline(
        file_path: Union[str, Path], content: str
    ) -> dict[str, Any]:
        file_path = Path(file_path)
        sync_result = materialize_source_document_to_chunks(
            source_path=file_path,
            content=content,
            datasource=resolved_datasource,
        )
        records_cleared = purge_extracted_records_for_doc_name(
            resolved_datasource, sync_result["doc_dir_name"]
        )
        revive_pipeline_output(sync_result["doc_dir_name"], resolved_datasource)
        revive_pipeline_output(file_path.stem, resolved_datasource)
        revive_pipeline_output(file_path.name, resolved_datasource)
        return build_pipeline_sync_payload(
            sync_result=sync_result,
            records_cleared=records_cleared,
        )

    async def index_file(file_path: Union[str, Path]) -> None:
        """Index one file inside the folder with support for multiple file formats."""
        file_path = Path(file_path)
        content = await load_file_content(file_path)

        if content:
            doc_id = build_doc_id(content)
            doc_registry.mark_processing(file_path, doc_id)
            try:
                await index_rag.ainsert(content)
                doc_manager.mark_as_indexed(file_path)
                doc_registry.mark_indexed(
                    file_path, doc_id, await get_doc_status_snapshot(doc_id)
                )
                sync_source_document_pipeline(file_path, content)
                logging.info(f"Successfully indexed file: {file_path}")
            except Exception as e:
                doc_registry.mark_failed(file_path, str(e), doc_id=doc_id)
                raise
        else:
            doc_registry.mark_failed(file_path, "No content extracted from file")
            logging.warning(f"No content extracted from file: {file_path}")

    async def index_files(file_paths: List[Union[str, Path]]) -> int:
        """Index a batch of files so MiniRAG can parallelize work internally."""
        valid_file_paths: list[Path] = []
        contents: list[str] = []
        doc_ids: list[str] = []

        for raw_file_path in file_paths:
            file_path = Path(raw_file_path)
            try:
                content = await load_file_content(file_path)
            except Exception as e:
                doc_registry.mark_failed(file_path, str(e))
                logging.error(f"Error reading file {file_path}: {str(e)}")
                continue

            if not content:
                doc_registry.mark_failed(file_path, "No content extracted from file")
                logging.warning(f"No content extracted from file: {file_path}")
                continue

            doc_id = build_doc_id(content)
            doc_registry.mark_processing(file_path, doc_id)
            valid_file_paths.append(file_path)
            contents.append(content)
            doc_ids.append(doc_id)

        if not contents:
            return 0

        try:
            await index_rag.ainsert(contents)
        except Exception as e:
            for file_path, doc_id in zip(valid_file_paths, doc_ids):
                doc_registry.mark_failed(file_path, str(e), doc_id=doc_id)
            raise

        for file_path, doc_id, content in zip(valid_file_paths, doc_ids, contents):
            doc_manager.mark_as_indexed(file_path)
            doc_registry.mark_indexed(
                file_path, doc_id, await get_doc_status_snapshot(doc_id)
            )
            sync_source_document_pipeline(file_path, content)
            logging.info(f"Successfully indexed file: {file_path}")
        return len(valid_file_paths)

    def _normalize_inventory_lookup(value: Optional[str]) -> Optional[str]:
        normalized = str(value or "").strip()
        if not normalized:
            return None

        candidate = Path(normalized)
        if candidate.is_absolute():
            return str(candidate.resolve())
        return str(candidate)

    def _document_inventory_aliases(
        item: Dict[str, Any], datasource: ResolvedDatasource
    ) -> set[str]:
        aliases: set[str] = set()
        for raw_value in (item.get("relative_path"), item.get("absolute_path")):
            normalized_value = _normalize_inventory_lookup(raw_value)
            if normalized_value:
                aliases.add(normalized_value)

        absolute_path = item.get("absolute_path")
        if not absolute_path:
            return aliases

        resolved_path = Path(absolute_path).resolve()
        for root in (
            datasource.source_root,
            datasource.staging_root,
            datasource.output_root,
            datasource.root,
            REPO_ROOT,
        ):
            try:
                aliases.add(str(resolved_path.relative_to(root)))
            except ValueError:
                continue

        return aliases

    def resolve_document_inventory_item(
        relative_path: str, datasource: ResolvedDatasource
    ) -> Optional[Dict[str, Any]]:
        lookup_path = _normalize_inventory_lookup(relative_path)
        if not lookup_path:
            return None

        doc_registry.sync_from_disk(doc_manager.supported_extensions)
        doc_manager.indexed_files = doc_registry.get_indexed_paths()
        summary = doc_registry.build_summary()
        for item in summary["items"]:
            if lookup_path in _document_inventory_aliases(item, datasource):
                return item

        pipeline_summary = build_pipeline_outputs_summary(datasource)
        for item in pipeline_summary["items"]:
            if lookup_path in _document_inventory_aliases(item, datasource):
                return item

        return None

    def require_mutable_document_item(
        relative_path: str,
        datasource: ResolvedDatasource,
        operation: str,
    ) -> Dict[str, Any]:
        item = resolve_document_inventory_item(relative_path, datasource)
        if item is None:
            raise HTTPException(status_code=404, detail="Document not found")
        if item.get("source_kind") == "pipeline":
            raise HTTPException(
                status_code=400,
                detail=f"pipeline resources do not support {operation}",
            )
        return item

    @app.post("/documents/scan", dependencies=[Depends(optional_api_key)])
    async def scan_for_new_documents():
        """Trigger the scanning process"""
        global scan_progress

        try:
            with progress_lock:
                if scan_progress["is_scanning"]:
                    return {"status": "already_scanning"}

                scan_progress["is_scanning"] = True
                scan_progress["indexed_count"] = 0
                scan_progress["progress"] = 0

            new_files = doc_manager.scan_directory_for_new_files()
            scan_progress["total_files"] = len(new_files)

            for i in range(0, len(new_files), args.batch_insert_size):
                batch_files = new_files[i : i + args.batch_insert_size]
                try:
                    with progress_lock:
                        batch_label = Path(batch_files[0]).name
                        if len(batch_files) > 1:
                            batch_label = (
                                f"{batch_label} (+{len(batch_files) - 1} more)"
                            )
                        scan_progress["current_file"] = batch_label

                    indexed_count = await index_files(batch_files)

                    with progress_lock:
                        scan_progress["indexed_count"] += indexed_count
                        scan_progress["progress"] = (
                            scan_progress["indexed_count"]
                            / scan_progress["total_files"]
                        ) * 100

                except Exception as e:
                    logging.error(
                        f"Error indexing files {[str(path) for path in batch_files]}: {str(e)}"
                    )

            return {
                "status": "success",
                "indexed_count": scan_progress["indexed_count"],
                "total_documents": len(doc_manager.indexed_files),
            }
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))
        finally:
            with progress_lock:
                scan_progress["is_scanning"] = False

    @app.get("/documents/scan-progress")
    async def get_scan_progress():
        """Get the current scanning progress"""
        with progress_lock:
            return scan_progress

    @app.get("/documents/summary", dependencies=[Depends(optional_api_key)])
    async def get_documents_summary(
        datasource_id: Optional[str] = Query(None),
        datasource_root: Optional[str] = Query(None),
        input_dir: Optional[str] = Query(None),
        working_dir: Optional[str] = Query(None),
        outputs_root: Optional[str] = Query(None),
    ):
        request_datasource = resolve_graph_datasource_request(
            datasource_id=datasource_id,
            datasource_root=datasource_root,
            input_dir=input_dir,
            working_dir=working_dir,
            outputs_root=outputs_root,
        )
        doc_registry.sync_from_disk(doc_manager.supported_extensions)
        doc_manager.indexed_files = doc_registry.get_indexed_paths()
        summary = doc_registry.build_summary()
        pipeline_summary = build_pipeline_outputs_summary(request_datasource)

        live_status_items: List[Dict[str, Any]] = []
        for item in summary["items"]:
            if item.get("source_kind") == "pipeline" or not item.get("doc_id"):
                live_status_items.append(item)
                continue
            doc_status = await get_doc_status_snapshot(item["doc_id"])
            live_status_items.append(
                merge_inventory_item_with_doc_status(item, doc_status)
            )

        live_status_items = await asyncio.gather(
            *(enrich_inventory_item_fallback_summary(item) for item in live_status_items)
        )
        live_status_items = [
            merge_source_item_with_pipeline_graph_counts(
                item,
                find_matching_pipeline_item(item, pipeline_summary["items"])
                if item.get("source_kind", item.get("resource_kind", "input_dir"))
                == "input_dir"
                else None,
            )
            for item in live_status_items
        ]

        combined_items = list(live_status_items)
        existing_paths = {item.get("absolute_path") for item in combined_items}
        for item in pipeline_summary["items"]:
            if item.get("absolute_path") not in existing_paths:
                combined_items.append(item)

        combined_items.sort(key=lambda item: item.get("modified_at", ""), reverse=True)
        stats = {
            "total": len(combined_items),
            "indexed": sum(
                1 for item in combined_items if item.get("status") == "indexed"
            ),
            "processing": sum(
                1 for item in combined_items if item.get("status") == "processing"
            ),
            "failed": sum(
                1 for item in combined_items if item.get("status") == "failed"
            ),
            "pending": sum(
                1 for item in combined_items if item.get("status") == "pending"
            ),
        }

        return {
            "datasource_id": request_datasource.id,
            "datasource": build_datasource_scope(request_datasource),
            "supported_extensions": list(doc_manager.supported_extensions),
            "stats": stats,
            "items": combined_items,
        }

    @app.post("/documents/upload", dependencies=[Depends(optional_api_key)])
    async def upload_to_input_dir(file: UploadFile = File(...)):
        """
        Endpoint for uploading a file to the input directory and indexing it.

        This API endpoint accepts a file through an HTTP POST request, checks if the
        uploaded file is of a supported type, saves it in the specified input directory,
        indexes it for retrieval, and returns a success status with relevant details.

        Parameters:
            file (UploadFile): The file to be uploaded. It must have an allowed extension as per
                               `doc_manager.supported_extensions`.

        Returns:
            dict: A dictionary containing the upload status ("success"),
                  a message detailing the operation result, and
                  the total number of indexed documents.

        Raises:
            HTTPException: If the file type is not supported, it raises a 400 Bad Request error.
                           If any other exception occurs during the file handling or indexing,
                           it raises a 500 Internal Server Error with details about the exception.
        """
        try:
            file_path = doc_manager.resolve_upload_target(file.filename)
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)

            # Immediately index the uploaded file
            await index_file(file_path)

            return {
                "status": "success",
                "message": f"File uploaded and indexed: {file_path.name}",
                "total_documents": len(doc_manager.indexed_files),
            }
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post(
        "/query", response_model=QueryResponse, dependencies=[Depends(optional_api_key)]
    )
    async def query_text(request: QueryRequest):
        """
        Handle a POST request at the /query endpoint to process user queries using RAG capabilities.

        Parameters:
            request (QueryRequest): A Pydantic model containing the following fields:
                - query (str): The text of the user's query.
                - mode (ModeEnum): Optional. Specifies the mode of retrieval augmentation.
                - stream (bool): Optional. Determines if the response should be streamed.
                - only_need_context (bool): Optional. If true, returns only the context without further processing.

        Returns:
            QueryResponse: A Pydantic model containing the result of the query processing.
                           If a string is returned (e.g., cache hit), it's directly returned.
                           Otherwise, an async generator may be used to build the response.

        Raises:
            HTTPException: Raised when an error occurs during the request handling process,
                           with status code 500 and detail containing the exception message.
        """
        try:
            response = await execute_query_request(
                request, provider_stream=request.stream
            )

            if isinstance(response, str):
                return QueryResponse(
                    response=ensure_non_empty_query_text(response)
                )

            result = await collect_query_response_text(response)
            return QueryResponse(response=result)
        except EmptyQueryResponseError as e:
            raise HTTPException(status_code=502, detail=str(e))
        except Exception as e:
            trace_exception(e)
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/query/plain", dependencies=[Depends(optional_api_key)])
    async def query_text_plain(request: QueryRequest):
        """
        Return plain text only, without the JSON response wrapper.
        Useful for lightweight frontends that only need final answer text.
        """
        try:
            response = await execute_query_request(request, provider_stream=False)
            result = await collect_query_response_text(response)

            from fastapi.responses import PlainTextResponse

            return PlainTextResponse(result, media_type="text/plain; charset=utf-8")
        except EmptyQueryResponseError as e:
            raise HTTPException(status_code=502, detail=str(e))
        except Exception as e:
            trace_exception(e)
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/query/stream", dependencies=[Depends(optional_api_key)])
    async def query_text_stream(request: QueryRequest):
        """
        This endpoint performs a retrieval-augmented generation (RAG) query and streams the response.

        Args:
            request (QueryRequest): The request object containing the query parameters.
            optional_api_key (Optional[str], optional): An optional API key for authentication. Defaults to None.

        Returns:
            StreamingResponse: A streaming response containing the RAG query results.
        """
        try:
            response = await execute_query_request(request, provider_stream=True)

            from fastapi.responses import StreamingResponse

            async def stream_generator():
                if isinstance(response, str):
                    yield f"{json.dumps({'response': ensure_non_empty_query_text(response)})}\n"
                else:
                    try:
                        async for chunk in response:
                            if chunk:
                                yield f"{json.dumps({'response': chunk})}\n"
                    except EmptyQueryResponseError as e:
                        yield f"{json.dumps({'error': str(e)})}\n"
                    except Exception as e:
                        logging.error(f"Streaming error: {str(e)}")
                        yield f"{json.dumps({'error': str(e)})}\n"

            return StreamingResponse(
                stream_generator(),
                media_type="application/x-ndjson",
                headers={
                    "Cache-Control": "no-cache",
                    "Connection": "keep-alive",
                    "Content-Type": "application/x-ndjson",
                    "Access-Control-Allow-Origin": "*",
                    "Access-Control-Allow-Methods": "POST, OPTIONS",
                    "Access-Control-Allow-Headers": "Content-Type",
                    "X-Accel-Buffering": "no",  # Disable Nginx buffering
                },
            )
        except EmptyQueryResponseError as e:
            raise HTTPException(status_code=502, detail=str(e))
        except Exception as e:
            trace_exception(e)
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/query/stream/plain", dependencies=[Depends(optional_api_key)])
    async def query_text_stream_plain(request: QueryRequest):
        """
        Stream plain text only, without NDJSON wrappers.
        Useful when the frontend wants raw incremental answer text.
        """
        try:
            response = await execute_query_request(request, provider_stream=True)

            if isinstance(response, str):
                async def single_chunk_generator():
                    yield ensure_non_empty_query_text(response)

                return build_plain_streaming_response(single_chunk_generator())

            first_chunk = ""
            try:
                while True:
                    try:
                        candidate = await response.__anext__()
                    except StopAsyncIteration:
                        break
                    if candidate:
                        first_chunk = str(candidate)
                        break
            except Exception as e:
                logging.error(f"Plain streaming preflight error: {str(e)}")
                raise

            if not first_chunk:
                fallback_response = await execute_query_request(
                    request, provider_stream=False
                )
                fallback_text = await collect_query_response_text(fallback_response)

                async def fallback_generator():
                    yield fallback_text

                return build_plain_streaming_response(fallback_generator())

            async def stream_generator():
                yield first_chunk
                try:
                    async for chunk in response:
                        if chunk:
                            yield str(chunk)
                except Exception as e:
                    logging.error(f"Plain streaming error: {str(e)}")

            return build_plain_streaming_response(stream_generator())
        except EmptyQueryResponseError as e:
            raise HTTPException(status_code=502, detail=str(e))
        except Exception as e:
            trace_exception(e)
            raise HTTPException(status_code=500, detail=str(e))

    @app.post(
        "/documents/text",
        response_model=InsertResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def insert_text(request: InsertTextRequest):
        """
        Insert text into the Retrieval-Augmented Generation (RAG) system.

        This endpoint allows you to insert text data into the RAG system for later retrieval and use in generating responses.

        Args:
            request (InsertTextRequest): The request body containing the text to be inserted.

        Returns:
            InsertResponse: A response object containing the status of the operation, a message, and the number of documents inserted.
        """
        try:
            await index_rag.ainsert(request.text)
            return InsertResponse(
                status="success",
                message="Text successfully inserted",
                document_count=1,
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post(
        "/documents/file",
        response_model=InsertResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def insert_file(file: UploadFile = File(...), description: str = Form(None)):
        """Insert a file directly into the RAG system

        Args:
            file: Uploaded file
            description: Optional description of the file

        Returns:
            InsertResponse: Status of the insertion operation

        Raises:
            HTTPException: For unsupported file types or processing errors
        """
        try:
            content = ""
            # Get file extension in lowercase
            ext = Path(file.filename).suffix.lower()

            match ext:
                case ".txt" | ".md":
                    # Text files handling
                    text_content = await file.read()
                    content = text_content.decode("utf-8")

                case ".xlsx":
                    if not pm.is_installed("openpyxl"):
                        pm.install("openpyxl")
                    from io import BytesIO
                    from openpyxl import load_workbook

                    workbook = load_workbook(
                        filename=BytesIO(await file.read()),
                        data_only=True,
                    )
                    sheet_blocks: list[str] = []
                    for sheet in workbook.worksheets:
                        rows: list[str] = []
                        for row in sheet.iter_rows(values_only=True):
                            cells = [
                                str(value).strip()
                                for value in row
                                if value is not None and str(value).strip()
                            ]
                            if cells:
                                rows.append(" | ".join(cells))
                        if rows:
                            sheet_blocks.append(
                                f"# {sheet.title}\n" + "\n".join(rows)
                            )
                    content = "\n\n".join(sheet_blocks)

                case _:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Unsupported file type. Supported types: {doc_manager.supported_extensions}",
                    )

            # Insert content into RAG system
            if content:
                # Add description if provided
                if description:
                    content = f"{description}\n\n{content}"

                await index_rag.ainsert(content)
                logging.info(f"Successfully indexed file: {file.filename}")

                return InsertResponse(
                    status="success",
                    message=f"File '{file.filename}' successfully inserted",
                    document_count=1,
                )
            else:
                raise HTTPException(
                    status_code=400,
                    detail="No content could be extracted from the file",
                )

        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="File encoding not supported")
        except Exception as e:
            logging.error(f"Error processing file {file.filename}: {str(e)}")
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/documents/file", dependencies=[Depends(optional_api_key)])
    async def get_document_file_detail(
        relative_path: str = Query(...),
        datasource_id: Optional[str] = Query(None),
        datasource_root: Optional[str] = Query(None),
        input_dir: Optional[str] = Query(None),
        working_dir: Optional[str] = Query(None),
        outputs_root: Optional[str] = Query(None),
    ):
        request_datasource = resolve_graph_datasource_request(
            datasource_id=datasource_id,
            datasource_root=datasource_root,
            input_dir=input_dir,
            working_dir=working_dir,
            outputs_root=outputs_root,
        )
        item = resolve_document_inventory_item(relative_path, request_datasource)
        if item is None:
            raise HTTPException(status_code=404, detail="Document not found")

        file_path = Path(item["absolute_path"])
        detail: Dict[str, Any] = {
            "datasource_id": request_datasource.id,
            "capabilities": dict(
                INPUT_DIR_CAPABILITIES
                if item.get("source_kind", "input_dir") == "input_dir"
                else PIPELINE_CAPABILITIES
            ),
            "resource_kind": item.get("source_kind", "input_dir"),
            "registry_snapshot": item,
            "file_stats": {"exists": file_path.exists()},
        }
        if file_path.exists():
            stat = file_path.stat()
            detail["file_stats"].update(
                {
                    "size": stat.st_size,
                    "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                }
            )
        if item.get("source_kind") != "pipeline":
            doc_status = (
                await get_doc_status_snapshot(item["doc_id"])
                if item.get("doc_id")
                else None
            )
            detail["doc_status"] = doc_status
            registry_snapshot = item
            if item.get("doc_id"):
                registry_snapshot = merge_inventory_item_with_doc_status(
                    registry_snapshot, doc_status
                )
            registry_snapshot = await enrich_inventory_item_fallback_summary(
                registry_snapshot
            )
            detail["registry_snapshot"] = merge_source_item_with_pipeline_graph_counts(
                registry_snapshot,
                find_matching_pipeline_item(
                    registry_snapshot,
                    build_pipeline_outputs_summary(request_datasource)["items"],
                ),
            )
        else:
            detail["doc_status"] = None
            detail["registry_snapshot"] = await enrich_inventory_item_fallback_summary(
                item
            )
        return detail

    @app.post(
        "/documents/file/reindex",
        response_model=InsertResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def reindex_document(
        relative_path: str = Query(...),
        datasource_id: Optional[str] = Query(None),
        datasource_root: Optional[str] = Query(None),
        input_dir: Optional[str] = Query(None),
        working_dir: Optional[str] = Query(None),
        outputs_root: Optional[str] = Query(None),
    ):
        request_datasource = resolve_graph_datasource_request(
            datasource_id=datasource_id,
            datasource_root=datasource_root,
            input_dir=input_dir,
            working_dir=working_dir,
            outputs_root=outputs_root,
        )
        item = require_mutable_document_item(
            relative_path, request_datasource, "reindex"
        )

        file_path = Path(item["absolute_path"])
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Document file not found")

        content = await load_file_content(file_path)
        if not content:
            raise HTTPException(
                status_code=400, detail="No content extracted from file"
            )

        doc_id = build_doc_id(content)
        doc_registry.mark_processing(file_path, doc_id)
        try:
            await index_rag.areindex_document(
                content,
                new_doc_id=doc_id,
                purge_doc_id=item.get("doc_id"),
            )
        except Exception as e:
            doc_registry.mark_failed(file_path, str(e), doc_id=doc_id)
            raise
        doc_manager.mark_as_indexed(file_path)
        doc_registry.mark_indexed(
            file_path, doc_id, await get_doc_status_snapshot(doc_id)
        )
        sync_source_document_pipeline(file_path, content)
        return InsertResponse(
            status="success",
            message=f"Document reindexed: {file_path.name}",
            document_count=len(doc_manager.indexed_files),
        )

    @app.post(
        "/documents/file/reprocess",
        response_model=InsertResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def reprocess_document(
        relative_path: str = Query(...),
        datasource_id: Optional[str] = Query(None),
        datasource_root: Optional[str] = Query(None),
        input_dir: Optional[str] = Query(None),
        working_dir: Optional[str] = Query(None),
        outputs_root: Optional[str] = Query(None),
    ):
        request_datasource = resolve_graph_datasource_request(
            datasource_id=datasource_id,
            datasource_root=datasource_root,
            input_dir=input_dir,
            working_dir=working_dir,
            outputs_root=outputs_root,
        )
        item = require_mutable_document_item(
            relative_path, request_datasource, "reprocess"
        )

        file_path = Path(item["absolute_path"])
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Document file not found")

        content = await load_file_content(file_path)
        if not content:
            raise HTTPException(
                status_code=400, detail="No content extracted from file"
            )

        doc_id = build_doc_id(content)
        doc_registry.mark_processing(file_path, doc_id)
        try:
            await index_rag.areindex_document(
                content,
                new_doc_id=doc_id,
                purge_doc_id=item.get("doc_id"),
            )
        except Exception as e:
            doc_registry.mark_failed(file_path, str(e), doc_id=doc_id)
            raise
        doc_manager.mark_as_indexed(file_path)
        doc_registry.mark_indexed(
            file_path, doc_id, await get_doc_status_snapshot(doc_id)
        )
        sync_source_document_pipeline(file_path, content)
        return InsertResponse(
            status="success",
            message=f"Document reprocessed: {file_path.name}",
            document_count=len(doc_manager.indexed_files),
        )

    @app.post(
        "/documents/file/ready-to-query",
        response_model=InsertResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def make_document_ready_to_query(
        relative_path: str = Query(...),
        datasource_id: Optional[str] = Query(None),
        datasource_root: Optional[str] = Query(None),
        input_dir: Optional[str] = Query(None),
        working_dir: Optional[str] = Query(None),
        outputs_root: Optional[str] = Query(None),
    ):
        request_datasource = resolve_graph_datasource_request(
            datasource_id=datasource_id,
            datasource_root=datasource_root,
            input_dir=input_dir,
            working_dir=working_dir,
            outputs_root=outputs_root,
        )
        item = require_mutable_document_item(
            relative_path, request_datasource, "reprocess"
        )

        file_path = Path(item["absolute_path"])
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Document file not found")

        stop_event = threading.Event()
        pipeline_key = register_document_pipeline(
            request_datasource, file_path, stop_event
        )
        try:
            summary = await run_ready_to_query_pipeline(
                request_datasource=request_datasource,
                file_path=file_path,
                stop_event=stop_event,
            )
        except asyncio.CancelledError:
            stop_event.set()
            raise HTTPException(
                status_code=409, detail="Document processing cancelled"
            )
        except RuntimeError as exc:
            if "cancelled" in str(exc).lower():
                raise HTTPException(
                    status_code=409, detail="Document processing cancelled"
                )
            raise
        finally:
            unregister_document_pipeline(pipeline_key, stop_event)
        return InsertResponse(
            status="success",
            message=(
                f"Document ready to query: {file_path.name}; "
                f"fault records kept={summary['kept_record_count']}, "
                f"accepted={summary['accepted_count']}, "
                f"rejected={summary['rejected_count']}"
            ),
            document_count=len(doc_manager.indexed_files),
        )

    @app.delete(
        "/documents/file",
        response_model=InsertResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def delete_document_file(
        relative_path: str = Query(...),
        datasource_id: Optional[str] = Query(None),
        datasource_root: Optional[str] = Query(None),
        input_dir: Optional[str] = Query(None),
        working_dir: Optional[str] = Query(None),
        outputs_root: Optional[str] = Query(None),
    ):
        request_datasource = resolve_graph_datasource_request(
            datasource_id=datasource_id,
            datasource_root=datasource_root,
            input_dir=input_dir,
            working_dir=working_dir,
            outputs_root=outputs_root,
        )
        item = require_mutable_document_item(
            relative_path, request_datasource, "delete"
        )

        file_path = Path(item["absolute_path"])
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Document file not found")

        request_document_pipeline_stop(request_datasource, file_path)
        content = await load_file_content(file_path)
        doc_id = item.get("doc_id") or (build_doc_id(content) if content else None)
        if not doc_id:
            raise HTTPException(status_code=404, detail="Document not found")

        file_path.unlink()
        await index_rag.adelete_document(doc_id)
        purge_pipeline_outputs_for_source_file(file_path, request_datasource)
        doc_registry.delete(file_path)
        doc_manager.indexed_files.discard(file_path)

        return InsertResponse(
            status="success",
            message=f"Document deleted: {file_path.name}",
            document_count=len(doc_manager.indexed_files),
        )

    @app.post(
        "/documents/batch",
        response_model=InsertResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def insert_batch(files: List[UploadFile] = File(...)):
        """Process multiple files in batch mode

        Args:
            files: List of files to process

        Returns:
            InsertResponse: Status of the batch insertion operation

        Raises:
            HTTPException: For processing errors
        """
        try:
            inserted_count = 0
            failed_files = []

            for file in files:
                try:
                    content = ""
                    ext = Path(file.filename).suffix.lower()

                    match ext:
                        case ".txt" | ".md":
                            text_content = await file.read()
                            content = text_content.decode("utf-8")

                        case ".xlsx":
                            if not pm.is_installed("openpyxl"):
                                pm.install("openpyxl")
                            from io import BytesIO
                            from openpyxl import load_workbook

                            workbook = load_workbook(
                                filename=BytesIO(await file.read()),
                                data_only=True,
                            )
                            sheet_blocks: list[str] = []
                            for sheet in workbook.worksheets:
                                rows: list[str] = []
                                for row in sheet.iter_rows(values_only=True):
                                    cells = [
                                        str(value).strip()
                                        for value in row
                                        if value is not None and str(value).strip()
                                    ]
                                    if cells:
                                        rows.append(" | ".join(cells))
                                if rows:
                                    sheet_blocks.append(
                                        f"# {sheet.title}\n" + "\n".join(rows)
                                    )
                            content = "\n\n".join(sheet_blocks)

                        case _:
                            failed_files.append(f"{file.filename} (unsupported type)")
                            continue

                    if content:
                        await index_rag.ainsert(content)
                        inserted_count += 1
                        logging.info(f"Successfully indexed file: {file.filename}")
                    else:
                        failed_files.append(f"{file.filename} (no content extracted)")

                except UnicodeDecodeError:
                    failed_files.append(f"{file.filename} (encoding error)")
                except Exception as e:
                    failed_files.append(f"{file.filename} ({str(e)})")
                    logging.error(f"Error processing file {file.filename}: {str(e)}")

            # Prepare status message
            if inserted_count == len(files):
                status = "success"
                status_message = f"Successfully inserted all {inserted_count} documents"
            elif inserted_count > 0:
                status = "partial_success"
                status_message = f"Successfully inserted {inserted_count} out of {len(files)} documents"
                if failed_files:
                    status_message += f". Failed files: {', '.join(failed_files)}"
            else:
                status = "failure"
                status_message = "No documents were successfully inserted"
                if failed_files:
                    status_message += f". Failed files: {', '.join(failed_files)}"

            return InsertResponse(
                status=status,
                message=status_message,
                document_count=inserted_count,
            )

        except Exception as e:
            logging.error(f"Batch processing error: {str(e)}")
            raise HTTPException(status_code=500, detail=str(e))

    @app.delete(
        "/documents",
        response_model=InsertResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def clear_documents():
        """
        Clear all documents from the MiniRAG system.

        This endpoint deletes all text chunks, entities vector database, and relationships vector database,
        effectively clearing all documents from the MiniRAG system.

        Returns:
            InsertResponse: A response object containing the status, message, and the new document count (0 in this case).
        """
        try:
            index_rag.text_chunks = []
            query_rag.text_chunks = []
            index_rag.entities_vdb = None
            query_rag.entities_vdb = None
            index_rag.relationships_vdb = None
            query_rag.relationships_vdb = None
            return InsertResponse(
                status="success",
                message="All documents cleared successfully",
                document_count=0,
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    # query all graph labels
    @app.get("/graph/label/list", dependencies=[Depends(optional_api_key)])
    async def get_graph_labels(
        datasource_id: str = Query(...),
        limit: int = Query(default=0, ge=0),
    ):
        scope_id = require_graph_datasource(datasource_id)
        return await query_rag.get_graph_labels(
            datasource_id=scope_id,
            limit=limit or None,
        )

    @app.get("/graph/label/entries", dependencies=[Depends(optional_api_key)])
    async def get_graph_label_entries(
        datasource_id: str = Query(...),
        limit: int = Query(default=0, ge=0),
    ):
        scope_id = require_graph_datasource(datasource_id)
        return await query_rag.get_graph_label_entries(
            datasource_id=scope_id,
            limit=limit or None,
        )

    # query all graph
    @app.get("/graphs", dependencies=[Depends(optional_api_key)])
    async def get_graphs(
        datasource_id: str = Query(...),
        mode: str = Query(default="label"),
        label: Optional[str] = Query(default=None),
        max_depth: int = Query(default=100, ge=0),
        max_nodes: int = Query(default=1000, ge=1),
        max_edges: int = Query(default=5000, ge=1),
    ):
        scope_id = require_graph_datasource(datasource_id)
        graph_mode = mode.strip().lower()

        if graph_mode == "full":
            try:
                graph = await query_rag.get_graph_full(
                    datasource_id=scope_id,
                    max_nodes=max_nodes,
                    max_edges=max_edges,
                )
            except ValueError as exc:
                detail = str(exc)
                if "too large" in detail.lower():
                    raise HTTPException(status_code=413, detail=detail)
                raise HTTPException(status_code=400, detail=detail)

            return {
                "datasource_id": scope_id,
                "datasource": graph_datasource_scope,
                "graph_mode": "full",
                "graph_state": "empty"
                if not graph["nodes"] and not graph["edges"]
                else "ready",
                **graph,
            }

        if not label:
            raise HTTPException(
                status_code=400, detail="label is required when mode is not full"
            )

        try:
            graph = await query_rag.get_graps(
                nodel_label=label,
                max_depth=max_depth,
                datasource_id=scope_id,
            )
        except ValueError as exc:
            raise HTTPException(status_code=404, detail=str(exc))

        return {
            "datasource_id": scope_id,
            "datasource": graph_datasource_scope,
            "graph_mode": "label",
            "label": label,
            **graph,
        }

    @app.get("/graph/summary", dependencies=[Depends(optional_api_key)])
    async def get_graph_summary(datasource_id: str = Query(...)):
        scope_id = require_graph_datasource(datasource_id)
        summary = await query_rag.get_graph_summary(datasource_id=scope_id)
        return {
            "datasource_id": scope_id,
            "datasource": graph_datasource_scope,
            **summary,
        }

    @app.get("/graph/node-detail", dependencies=[Depends(optional_api_key)])
    async def get_graph_node_detail(
        datasource_id: str = Query(...),
        label: str = Query(...),
        max_relationships: int = Query(default=20, ge=1),
    ):
        scope_id = require_graph_datasource(datasource_id)
        detail = await query_rag.get_graph_node_detail(
            label,
            max_relationships=max_relationships,
            datasource_id=scope_id,
        )
        if not detail:
            raise HTTPException(status_code=404, detail="Graph node not found")
        return {
            "datasource_id": scope_id,
            "datasource": graph_datasource_scope,
            **detail,
        }

    @app.get("/system/capabilities", dependencies=[Depends(optional_api_key)])
    async def get_system_capabilities():
        alias_stats = alias_store.get_stats()
        active_models = get_active_model_configs()
        return {
            "supported_modes": [mode.value for mode in SearchMode],
            "default_mode": SearchMode.graph_text_hybrid.value,
            "recommended_mode": SearchMode.graph_text_hybrid.value,
            "current_demo_mainline": SearchMode.graph_text_hybrid.value,
            "supports_stream": True,
            "supports_only_need_context": True,
            "stream_protocol": "plain_text",
            "recommended_query_endpoint": "/query/stream/plain",
            "frontend_rules": [
                "前端必须显式传递 mode，不要依赖后端默认值",
                "推荐优先使用 graph_text_hybrid 作为默认问答模式",
                "graph_only 会围绕图谱节点与关联证据组织回答",
                "text_only 为旧模式兼容入口，不建议新前端继续使用",
                "keyword_search 是新的关键词检索模式，关键词以空格分隔并按 BM25 匹配",
                "RAG 问答页使用 /query/plain 与 /query/stream/plain",
                "图谱页面用于浏览节点、关系与结构信息",
                "documents 系列接口用于文档导入、扫描与状态查询",
            ],
            "top_k": args.top_k,
            "batch_insert_size": args.batch_insert_size,
            "max_tokens": args.max_tokens,
            "query_model": active_models["query"].model,
            "embedding_model": active_models["embedding"].model,
            "storage": {
                "graph": ollama_server_infos.GRAPH_STORAGE,
                "vector": ollama_server_infos.VECTOR_STORAGE,
                "doc_status": ollama_server_infos.DOC_STATUS_STORAGE,
            },
            "datasource_id": args.datasource_id,
            "datasource": graph_datasource_scope,
            "alias_store": alias_stats,
            "equipment_catalog": equipment_catalog,
        }

    class RuntimeModelSelectionUpdateRequest(BaseModel):
        index_llm_id: Optional[str] = None
        query_llm_id: Optional[str] = None
        embedding_model_id: Optional[str] = None
        unload_previous_query_model: bool = True

    class RuntimeModelSelectionResponse(RuntimeModelCatalogResponse):
        previous_query_unload: Optional[RuntimeUnloadResult] = None

    @app.get(
        "/system/models",
        response_model=RuntimeModelCatalogResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def get_system_models():
        return get_active_model_catalog()

    @app.post(
        "/system/models/select",
        response_model=RuntimeModelSelectionResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def select_system_models(request: RuntimeModelSelectionUpdateRequest):
        previous_query_model_config = model_registry.resolve_llm(
            active_model_selection.query_llm_id, role=LLM_ROLE_QUERY
        )
        try:
            next_selection = resolve_runtime_selection(
                model_registry,
                RuntimeModelSelection(
                    index_llm_id=request.index_llm_id
                    or active_model_selection.index_llm_id,
                    query_llm_id=request.query_llm_id
                    or active_model_selection.query_llm_id,
                    embedding_model_id=request.embedding_model_id
                    or active_model_selection.embedding_model_id,
                ),
            )
        except KeyError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

        rebuild_runtime_rags(next_selection)
        previous_query_unload: RuntimeUnloadResult | None = None
        next_index_model_config = model_registry.resolve_llm(
            active_model_selection.index_llm_id, role=LLM_ROLE_INDEX
        )
        next_query_model_config = model_registry.resolve_llm(
            active_model_selection.query_llm_id, role=LLM_ROLE_QUERY
        )
        previous_signature = _physical_model_signature(previous_query_model_config)
        still_in_use = previous_signature in {
            _physical_model_signature(next_index_model_config),
            _physical_model_signature(next_query_model_config),
        }

        if request.unload_previous_query_model and not still_in_use:
            previous_query_unload = await unload_lmstudio_model_instances(
                binding=previous_query_model_config.binding,
                host=previous_query_model_config.host,
                model=previous_query_model_config.model,
                api_key=previous_query_model_config.api_key,
            )
            if previous_query_unload.status == "failed":
                logging.warning(
                    "Failed to unload previous query model %s at %s: %s",
                    previous_query_model_config.model,
                    previous_query_model_config.host,
                    previous_query_unload.message,
                )
        elif request.unload_previous_query_model:
            previous_query_unload = RuntimeUnloadResult(
                status="skipped",
                binding=previous_query_model_config.binding,
                host=previous_query_model_config.host,
                model=previous_query_model_config.model,
                message="previous query model is still referenced by the active runtime",
            )

        catalog = get_active_model_catalog()
        return RuntimeModelSelectionResponse(
            **catalog.model_dump(),
            previous_query_unload=previous_query_unload,
        )

    @app.get("/system/config", dependencies=[Depends(optional_api_key)])
    async def get_system_config():
        alias_stats = alias_store.get_stats()
        available_modes = [mode.value for mode in SearchMode]
        default_mode = available_modes[0] if available_modes else "graph_text_hybrid"
        active_models = get_active_model_configs()
        return {
            "server": {
                "host": args.host,
                "port": args.port,
                "working_dir": str(args.working_dir),
                "input_dir": str(args.input_dir),
                "datasource_id": args.datasource_id,
                "datasource_root": graph_datasource_scope["datasource_root"],
                "source_root": graph_datasource_scope["source_root"],
                "staging_root": graph_datasource_scope["staging_root"],
                "output_root": graph_datasource_scope["output_root"],
                "log_level": args.log_level,
                "auto_scan_at_startup": bool(args.auto_scan_at_startup),
            },
            "llm": {
                "binding": active_models["index"].binding,
                "binding_host": active_models["index"].host,
                "model": active_models["index"].model,
                "model_id": active_model_selection.index_llm_id,
                "query_binding": active_models["query"].binding,
                "query_binding_host": active_models["query"].host,
                "query_model": active_models["query"].model,
                "query_model_id": active_model_selection.query_llm_id,
                "max_async": args.max_async,
                "max_tokens": args.max_tokens,
                "history_turns": args.history_turns,
            },
            "embedding": {
                "binding": active_models["embedding"].binding,
                "binding_host": active_models["embedding"].host,
                "model": active_models["embedding"].model,
                "model_id": active_model_selection.embedding_model_id,
                "dimension": args.embedding_dim,
                "max_embed_tokens": args.max_embed_tokens,
            },
            "chunking": {
                "chunk_size": args.chunk_size,
                "chunk_overlap_size": args.chunk_overlap_size,
                "batch_insert_size": args.batch_insert_size,
                "max_parallel_insert": args.max_parallel_insert,
            },
            "query": {
                "default_mode": default_mode,
                "available_modes": available_modes,
                "top_k": args.top_k,
                "cosine_threshold": args.cosine_threshold,
                "stream_protocol": "plain_text",
                "recommended_query_endpoint": "/query/stream/plain",
            },
            "storage": {
                "kv_storage": ollama_server_infos.KV_STORAGE,
                "doc_status_storage": ollama_server_infos.DOC_STATUS_STORAGE,
                "graph_storage": ollama_server_infos.GRAPH_STORAGE,
                "vector_storage": ollama_server_infos.VECTOR_STORAGE,
            },
            "model_selection": active_model_selection.model_dump(),
            "alias_store": alias_stats,
            "notes": [
                "知识库页与数据源配置页直接连接当前后端服务。",
                f"当前默认问答模式为 {default_mode}。",
                "流式问答推荐使用 /query/stream/plain。",
                "documents 系列接口用于文档目录扫描、上传与状态同步。",
                f"当前图谱 datasource_id 为 {args.datasource_id}。",
                "模型切换通过 model registry 的标识符完成，不再依赖代码内硬编码分支。",
            ],
        }

    @app.get(
        "/chat/state",
        response_model=ChatSessionStateResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def get_chat_state(datasource_id: Optional[str] = Query(default=None)):
        require_alias_datasource(datasource_id)
        return chat_session_store.get_state()

    @app.put(
        "/chat/state",
        response_model=ChatSessionStateResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def put_chat_state(request: ChatSessionStateRequest):
        require_alias_datasource(request.datasource_id)
        return chat_session_store.replace_state(
            request.sessions, request.active_session_id
        )

    @app.post(
        "/chat/export",
        response_model=ChatExportResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def export_chat_markdown(request: ChatExportRequest):
        require_alias_datasource(request.datasource_id)
        try:
            return chat_session_store.export_markdown(
                request.session_id, request.message_id
            )
        except KeyError as exc:
            detail = exc.args[0] if exc.args else "chat session not found"
            raise HTTPException(status_code=404, detail=str(detail)) from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post(
        "/chat/export/download",
        dependencies=[Depends(optional_api_key)],
    )
    async def export_chat_markdown_download(request: ChatExportRequest):
        require_alias_datasource(request.datasource_id)
        try:
            payload = chat_session_store.export_markdown(
                request.session_id, request.message_id
            )
            file_name = str(payload.get("file_name") or "chat-export.md")
            markdown = str(payload.get("content") or "")
            quoted_file_name = quote(file_name)
            return Response(
                content=markdown.encode("utf-8"),
                media_type="text/markdown; charset=utf-8",
                headers={
                    "Content-Disposition": (
                        f"attachment; filename*=UTF-8''{quoted_file_name}"
                    )
                },
            )
        except KeyError as exc:
            detail = exc.args[0] if exc.args else "chat session not found"
            raise HTTPException(status_code=404, detail=str(detail)) from exc
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get(
        "/aliases",
        response_model=AliasListResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def list_aliases(
        datasource_id: Optional[str] = Query(default=None),
        entity_type: Optional[str] = Query(default=None),
        enabled: Optional[bool] = Query(default=None),
        reviewed: Optional[bool] = Query(default=None),
        q: Optional[str] = Query(default=None),
    ):
        scope_id = require_alias_datasource(datasource_id)
        normalized_entity_type = entity_type.upper() if entity_type else None
        if (
            normalized_entity_type
            and normalized_entity_type not in SUPPORTED_ALIAS_ENTITY_TYPES
        ):
            raise HTTPException(status_code=400, detail="unsupported entity_type")

        items = alias_store.list_aliases(
            datasource_id=scope_id,
            entity_type=normalized_entity_type,
            enabled=enabled,
            reviewed=reviewed,
            query=q,
        )
        return {
            "datasource_id": scope_id,
            "items": items,
            "total": len(items),
            "stats": alias_store.get_stats(),
        }

    @app.post(
        "/aliases",
        response_model=AliasRecordResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def create_alias(request: AliasCreateRequest):
        try:
            scope_id = require_alias_datasource(request.datasource_id)
            return alias_store.create_alias(
                datasource_id=scope_id,
                canonical_name=request.canonical_name,
                entity_type=request.entity_type,
                alias=request.alias,
                enabled=request.enabled,
                reviewed=request.reviewed,
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    @app.patch(
        "/aliases/{alias_id}",
        response_model=AliasRecordResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def update_alias(alias_id: str, request: AliasUpdateRequest):
        try:
            scope_id = require_alias_datasource(request.datasource_id)
            return alias_store.update_alias(
                alias_id,
                datasource_id=scope_id,
                canonical_name=request.canonical_name,
                entity_type=request.entity_type,
                alias=request.alias,
                enabled=request.enabled,
                reviewed=request.reviewed,
            )
        except KeyError:
            raise HTTPException(status_code=404, detail="alias not found")
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))

    @app.delete("/aliases/{alias_id}", dependencies=[Depends(optional_api_key)])
    async def delete_alias(
        alias_id: str,
        datasource_id: Optional[str] = Query(default=None),
    ):
        require_alias_datasource(datasource_id)
        try:
            alias_store.delete_alias(alias_id)
            return {"status": "ok"}
        except KeyError:
            raise HTTPException(status_code=404, detail="alias not found")

    @app.get(
        "/aliases/resolve",
        response_model=AliasResolveResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def resolve_alias_query(
        query: str,
        datasource_id: Optional[str] = Query(default=None),
    ):
        require_alias_datasource(datasource_id)
        return route_faultcase_query(query, alias_store)

    # Ollama compatible API endpoints
    # -------------------------------------------------
    @app.get("/api/version", dependencies=[Depends(optional_api_key)])
    async def get_version():
        """Get Ollama version information"""
        return OllamaVersionResponse(version="0.5.4")

    @app.get("/api/tags", dependencies=[Depends(optional_api_key)])
    async def get_tags():
        """Get available models"""
        return OllamaTagResponse(
            models=[
                {
                    "name": ollama_server_infos.LIGHTRAG_MODEL,
                    "model": ollama_server_infos.LIGHTRAG_MODEL,
                    "size": ollama_server_infos.LIGHTRAG_SIZE,
                    "digest": ollama_server_infos.LIGHTRAG_DIGEST,
                    "modified_at": ollama_server_infos.LIGHTRAG_CREATED_AT,
                    "details": {
                        "parent_model": "",
                        "format": "gguf",
                        "family": ollama_server_infos.LIGHTRAG_NAME,
                        "families": [ollama_server_infos.LIGHTRAG_NAME],
                        "parameter_size": "13B",
                        "quantization_level": "Q4_0",
                    },
                }
            ]
        )

    def parse_query_mode(query: str) -> tuple[str, SearchMode]:
        mode_map = {
            "/graph_text_hybrid ": SearchMode.graph_text_hybrid,
            "/graph_only ": SearchMode.graph_only,
            "/text_only ": SearchMode.text_only,
        }

        for prefix, mode in mode_map.items():
            if query.startswith(prefix):
                # After removing prefix an leading spaces
                cleaned_query = query[len(prefix) :].lstrip()
                return cleaned_query, mode

        return query, SearchMode.graph_text_hybrid

    @app.post("/api/generate", dependencies=[Depends(optional_api_key)])
    async def generate(raw_request: Request, request: OllamaGenerateRequest):
        """Handle generate completion requests
        For compatiblity purpuse, the request is not processed by MiniRAG,
        and will be handled by underlying LLM model.
        """
        try:
            query = request.prompt
            start_time = time.time_ns()
            prompt_tokens = estimate_tokens(query)

            if request.system:
                query_rag.llm_model_kwargs["system_prompt"] = request.system

            if request.stream:
                from fastapi.responses import StreamingResponse

                response = await query_rag.llm_model_func(
                    query, stream=True, **query_rag.llm_model_kwargs
                )

                async def stream_generator():
                    try:
                        first_chunk_time = None
                        last_chunk_time = None
                        total_response = ""

                        # Ensure response is an async generator
                        if isinstance(response, str):
                            # If it's a string, send in two parts
                            first_chunk_time = time.time_ns()
                            last_chunk_time = first_chunk_time
                            total_response = response

                            data = {
                                "model": ollama_server_infos.LIGHTRAG_MODEL,
                                "created_at": ollama_server_infos.LIGHTRAG_CREATED_AT,
                                "response": response,
                                "done": False,
                            }
                            yield f"{json.dumps(data, ensure_ascii=False)}\n"

                            completion_tokens = estimate_tokens(total_response)
                            total_time = last_chunk_time - start_time
                            prompt_eval_time = first_chunk_time - start_time
                            eval_time = last_chunk_time - first_chunk_time

                            data = {
                                "model": ollama_server_infos.LIGHTRAG_MODEL,
                                "created_at": ollama_server_infos.LIGHTRAG_CREATED_AT,
                                "done": True,
                                "total_duration": total_time,
                                "load_duration": 0,
                                "prompt_eval_count": prompt_tokens,
                                "prompt_eval_duration": prompt_eval_time,
                                "eval_count": completion_tokens,
                                "eval_duration": eval_time,
                            }
                            yield f"{json.dumps(data, ensure_ascii=False)}\n"
                        else:
                            async for chunk in response:
                                if chunk:
                                    if first_chunk_time is None:
                                        first_chunk_time = time.time_ns()

                                    last_chunk_time = time.time_ns()

                                    total_response += chunk
                                    data = {
                                        "model": ollama_server_infos.LIGHTRAG_MODEL,
                                        "created_at": ollama_server_infos.LIGHTRAG_CREATED_AT,
                                        "response": chunk,
                                        "done": False,
                                    }
                                    yield f"{json.dumps(data, ensure_ascii=False)}\n"

                            completion_tokens = estimate_tokens(total_response)
                            total_time = last_chunk_time - start_time
                            prompt_eval_time = first_chunk_time - start_time
                            eval_time = last_chunk_time - first_chunk_time

                            data = {
                                "model": ollama_server_infos.LIGHTRAG_MODEL,
                                "created_at": ollama_server_infos.LIGHTRAG_CREATED_AT,
                                "done": True,
                                "total_duration": total_time,
                                "load_duration": 0,
                                "prompt_eval_count": prompt_tokens,
                                "prompt_eval_duration": prompt_eval_time,
                                "eval_count": completion_tokens,
                                "eval_duration": eval_time,
                            }
                            yield f"{json.dumps(data, ensure_ascii=False)}\n"
                            return

                    except Exception as e:
                        logging.error(f"Error in stream_generator: {str(e)}")
                        raise

                return StreamingResponse(
                    stream_generator(),
                    media_type="application/x-ndjson",
                    headers={
                        "Cache-Control": "no-cache",
                        "Connection": "keep-alive",
                        "Content-Type": "application/x-ndjson",
                        "Access-Control-Allow-Origin": "*",
                        "Access-Control-Allow-Methods": "POST, OPTIONS",
                        "Access-Control-Allow-Headers": "Content-Type",
                    },
                )
            else:
                first_chunk_time = time.time_ns()
                response_text = await query_rag.llm_model_func(
                    query, stream=False, **query_rag.llm_model_kwargs
                )
                last_chunk_time = time.time_ns()

                if not response_text:
                    response_text = "No response generated"

                completion_tokens = estimate_tokens(str(response_text))
                total_time = last_chunk_time - start_time
                prompt_eval_time = first_chunk_time - start_time
                eval_time = last_chunk_time - first_chunk_time

                return {
                    "model": ollama_server_infos.LIGHTRAG_MODEL,
                    "created_at": ollama_server_infos.LIGHTRAG_CREATED_AT,
                    "response": str(response_text),
                    "done": True,
                    "total_duration": total_time,
                    "load_duration": 0,
                    "prompt_eval_count": prompt_tokens,
                    "prompt_eval_duration": prompt_eval_time,
                    "eval_count": completion_tokens,
                    "eval_duration": eval_time,
                }
        except Exception as e:
            trace_exception(e)
            raise HTTPException(status_code=500, detail=str(e))

    @app.post("/api/chat", dependencies=[Depends(optional_api_key)])
    async def chat(raw_request: Request, request: OllamaChatRequest):
        """Process chat completion requests.
        Routes user queries through MiniRAG by selecting query mode based on prefix indicators.
        Detects and forwards OpenWebUI session-related requests (for meta data generation task) directly to LLM.
        """
        try:
            # Get all messages
            messages = request.messages
            if not messages:
                raise HTTPException(status_code=400, detail="No messages provided")

            # Get the last message as query and previous messages as history
            query = messages[-1].content
            # Convert OllamaMessage objects to dictionaries
            # Temporarily disable backend multi-turn memory. Keep only the
            # current user query and ignore previous messages.
            # conversation_history = [
            #     {"role": msg.role, "content": msg.content} for msg in messages[:-1]
            # ]
            conversation_history = []

            # Check for query prefix
            cleaned_query, mode = parse_query_mode(query)

            start_time = time.time_ns()
            prompt_tokens = estimate_tokens(cleaned_query)

            param_dict = {
                "mode": mode,
                "stream": request.stream,
                "only_need_context": False,
                "conversation_history": conversation_history,
                "top_k": args.top_k,
            }

            # if args.history_turns is not None:
            #     param_dict["history_turns"] = args.history_turns
            param_dict["history_turns"] = 0

            query_param = QueryParam(**param_dict)

            if request.stream:
                from fastapi.responses import StreamingResponse

                response = await query_rag.aquery(  # Need await to get async generator
                    cleaned_query, param=query_param
                )

                async def stream_generator():
                    try:
                        first_chunk_time = None
                        last_chunk_time = None
                        total_response = ""

                        # Ensure response is an async generator
                        if isinstance(response, str):
                            # If it's a string, send in two parts
                            first_chunk_time = time.time_ns()
                            last_chunk_time = first_chunk_time
                            total_response = response

                            data = {
                                "model": ollama_server_infos.LIGHTRAG_MODEL,
                                "created_at": ollama_server_infos.LIGHTRAG_CREATED_AT,
                                "message": {
                                    "role": "assistant",
                                    "content": response,
                                    "images": None,
                                },
                                "done": False,
                            }
                            yield f"{json.dumps(data, ensure_ascii=False)}\n"

                            completion_tokens = estimate_tokens(total_response)
                            total_time = last_chunk_time - start_time
                            prompt_eval_time = first_chunk_time - start_time
                            eval_time = last_chunk_time - first_chunk_time

                            data = {
                                "model": ollama_server_infos.LIGHTRAG_MODEL,
                                "created_at": ollama_server_infos.LIGHTRAG_CREATED_AT,
                                "done": True,
                                "total_duration": total_time,
                                "load_duration": 0,
                                "prompt_eval_count": prompt_tokens,
                                "prompt_eval_duration": prompt_eval_time,
                                "eval_count": completion_tokens,
                                "eval_duration": eval_time,
                            }
                            yield f"{json.dumps(data, ensure_ascii=False)}\n"
                        else:
                            async for chunk in response:
                                if chunk:
                                    if first_chunk_time is None:
                                        first_chunk_time = time.time_ns()

                                    last_chunk_time = time.time_ns()

                                    total_response += chunk
                                    data = {
                                        "model": ollama_server_infos.LIGHTRAG_MODEL,
                                        "created_at": ollama_server_infos.LIGHTRAG_CREATED_AT,
                                        "message": {
                                            "role": "assistant",
                                            "content": chunk,
                                            "images": None,
                                        },
                                        "done": False,
                                    }
                                    yield f"{json.dumps(data, ensure_ascii=False)}\n"

                            completion_tokens = estimate_tokens(total_response)
                            total_time = last_chunk_time - start_time
                            prompt_eval_time = first_chunk_time - start_time
                            eval_time = last_chunk_time - first_chunk_time

                            data = {
                                "model": ollama_server_infos.LIGHTRAG_MODEL,
                                "created_at": ollama_server_infos.LIGHTRAG_CREATED_AT,
                                "done": True,
                                "total_duration": total_time,
                                "load_duration": 0,
                                "prompt_eval_count": prompt_tokens,
                                "prompt_eval_duration": prompt_eval_time,
                                "eval_count": completion_tokens,
                                "eval_duration": eval_time,
                            }
                            yield f"{json.dumps(data, ensure_ascii=False)}\n"
                            return  # Ensure the generator ends immediately after sending the completion marker
                    except Exception as e:
                        logging.error(f"Error in stream_generator: {str(e)}")
                        raise

                return StreamingResponse(
                    stream_generator(),
                    media_type="application/x-ndjson",
                    headers={
                        "Cache-Control": "no-cache",
                        "Connection": "keep-alive",
                        "Content-Type": "application/x-ndjson",
                        "Access-Control-Allow-Origin": "*",
                        "Access-Control-Allow-Methods": "POST, OPTIONS",
                        "Access-Control-Allow-Headers": "Content-Type",
                    },
                )
            else:
                first_chunk_time = time.time_ns()

                # Determine if the request is from Open WebUI's session title and session keyword generation task
                match_result = re.search(
                    r"\n<chat_history>\nUSER:", cleaned_query, re.MULTILINE
                )
                if match_result:
                    if request.system:
                        query_rag.llm_model_kwargs["system_prompt"] = request.system

                    response_text = await query_rag.llm_model_func(
                        cleaned_query, stream=False, **query_rag.llm_model_kwargs
                    )
                else:
                    response_text = await query_rag.aquery(
                        cleaned_query, param=query_param
                    )

                last_chunk_time = time.time_ns()

                if not response_text:
                    response_text = "No response generated"

                completion_tokens = estimate_tokens(str(response_text))
                total_time = last_chunk_time - start_time
                prompt_eval_time = first_chunk_time - start_time
                eval_time = last_chunk_time - first_chunk_time

                return {
                    "model": ollama_server_infos.LIGHTRAG_MODEL,
                    "created_at": ollama_server_infos.LIGHTRAG_CREATED_AT,
                    "message": {
                        "role": "assistant",
                        "content": str(response_text),
                        "images": None,
                    },
                    "done": True,
                    "total_duration": total_time,
                    "load_duration": 0,
                    "prompt_eval_count": prompt_tokens,
                    "prompt_eval_duration": prompt_eval_time,
                    "eval_count": completion_tokens,
                    "eval_duration": eval_time,
                }
        except Exception as e:
            trace_exception(e)
            raise HTTPException(status_code=500, detail=str(e))

    @app.get("/documents", dependencies=[Depends(optional_api_key)])
    async def documents():
        """Get current system status"""
        doc_registry.sync_from_disk(doc_manager.supported_extensions)
        doc_manager.indexed_files = doc_registry.get_indexed_paths()
        return [str(path) for path in sorted(doc_manager.indexed_files)]

    @app.get("/health", dependencies=[Depends(optional_api_key)])
    async def get_status():
        """Get current system status"""
        doc_registry.sync_from_disk(doc_manager.supported_extensions)
        doc_manager.indexed_files = doc_registry.get_indexed_paths()
        files = doc_manager.scan_directory()
        active_models = get_active_model_configs()
        return {
            "status": "healthy",
            "working_directory": str(args.working_dir),
            "input_directory": str(args.input_dir),
            "datasource": graph_datasource_scope,
            "indexed_files": [str(f) for f in files],
            "indexed_files_count": len(files),
            "configuration": {
                # LLM configuration binding/host address (if applicable)/model (if applicable)
                "llm_binding": active_models["index"].binding,
                "llm_binding_host": active_models["index"].host,
                "llm_model": active_models["index"].model,
                "llm_model_id": active_model_selection.index_llm_id,
                "query_llm_binding": active_models["query"].binding,
                "query_llm_binding_host": active_models["query"].host,
                "query_llm_model": active_models["query"].model,
                "query_llm_model_id": active_model_selection.query_llm_id,
                # embedding model configuration binding/host address (if applicable)/model (if applicable)
                "embedding_binding": active_models["embedding"].binding,
                "embedding_binding_host": active_models["embedding"].host,
                "embedding_model": active_models["embedding"].model,
                "embedding_model_id": active_model_selection.embedding_model_id,
                "max_tokens": args.max_tokens,
                "kv_storage": ollama_server_infos.KV_STORAGE,
                "doc_status_storage": ollama_server_infos.DOC_STATUS_STORAGE,
                "graph_storage": ollama_server_infos.GRAPH_STORAGE,
                "vector_storage": ollama_server_infos.VECTOR_STORAGE,
            },
        }

    # ============================================================
    # Benchmark REST Endpoints (T4.1 - T4.4)
    # ============================================================

    # Request models for benchmark endpoints
    # Request models for benchmark endpoints
    class BenchmarkRunRequest(BaseModel):
        selected_model: str = "qwen3.5-2b"
        judge_model_type: str = "cloud" # "cloud" or "local"

    class BenchmarkRunResponse(BaseModel):
        status: str
        run_id: str
        state: BenchmarkState
        selected_model: str
        message: str

    class BenchmarkStopRequest(BaseModel):
        reason: Optional[str] = None

    class BenchmarkStopResponse(BaseModel):
        status: str
        state: BenchmarkState
        message: str

    class BenchmarkResetResponse(BaseModel):
        status: str
        state: BenchmarkState
        message: str

    @app.post(
        "/benchmark/run",
        response_model=BenchmarkRunResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def start_benchmark_run(request: BenchmarkRunRequest):
        """Start a new benchmark run - requires judge for scoring."""
        try:
            selected_model_id = active_model_selection.query_llm_id
            benchmark_state.set_model(selected_model_id)
            run_id = f"run-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
            rows = ship_benchmark_adapter.load_source_rows()
            benchmark_model_config = model_registry.resolve_llm(
                selected_model_id, role=LLM_ROLE_QUERY
            )
            benchmark_embedding_config = model_registry.resolve_embedding(
                active_model_selection.embedding_model_id
            )

            benchmark_call_timeout = args.timeout or 30
            benchmark_judge_timeout = max(
                benchmark_call_timeout,
                int(os.getenv("BENCHMARK_JUDGE_TIMEOUT_SECONDS", "90")),
            )
            shared_query_profile = build_shared_query_profile()
            runtime_note = (
                f" (binding={benchmark_model_config.binding}, model={benchmark_model_config.model})"
            )
            result_timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            result_path = (
                ship_benchmark_adapter.source_csv_path.parent
                / f"query_set_judged_{result_timestamp}.csv"
            )
            result_writer = IncrementalBenchmarkCsvWriter(
                result_path,
                BENCHMARK_QUERY_MODES,
            )

            def build_benchmark_rag_instance():
                try:
                    return build_rag(
                        benchmark_model_config,
                        benchmark_embedding_config,
                    )
                except Exception as build_err:
                    raise RuntimeError(
                        f"Benchmark execution failed: could not build query runtime: {build_err}"
                    ) from build_err

            def judge_answers(
                row: dict[str, Any], mode_answers: dict[str, str], judge_model_type: str = "cloud"
            ) -> dict[str, int]:
                import requests
                
                judge_api_key = os.getenv("DEEPSEEK_API_KEY")
                if judge_model_type == "cloud" and not judge_api_key:
                    raise RuntimeError(
                        "DEEPSEEK_API_KEY is not set; live benchmark scoring cannot judge successful rows."
                    )

                output_schema = json.dumps(
                    {mode: 1 for mode in mode_answers}, ensure_ascii=False, indent=2
                )
                answers_block = "\n".join(
                    f"{mode}: {answer}" for mode, answer in mode_answers.items()
                )
                prompt = f"""
你是一位专业的 RAG 系统评测裁判。
现在，我会给你提供：一个问题、一份标准答案 (Gold Answer)、可选的参考证据，以及由不同 RAG 系统生成的几个回答。

评分规则（必须严格遵守）：
- **1 分 (正确)**：回答内容正确，且涵盖了标准答案的核心要点。即使措辞不同或包含了额外的正确细节，只要核心意思对齐，就应给 1 分。
- **0 分 (回避或不足/拒绝回答)**：回答与问题无关、回答不完整，或者 RAG 系统明确表示“无法确定”、“证据不足”、“我不知道”等拒绝回答的情况。**只要 RAG 系统输出类似于“根据当前证据/上下文无法回答”这种表态，必须打 0 分，严禁打 -1 分。**
- **-1 分 (错误)**：回答提供了标准答案中没有的、且在事实上错误或与标准答案直接冲突的信息。**不要因为系统拒绝回答而打 -1 分。**

请仅返回一个 JSON 对象。
使用下面提供的系统名称 (mode name) 作为 JSON 的键。

问题:
{row.get("question", "")}

标准答案:
{row.get("answer") or row.get("gold_answer") or ""}

参考证据:
{row.get("evidence", "")}

题目类型:
{row.get("question_type", "")}

系统待评价回答:
{answers_block}

预期返回 JSON 格式:
{output_schema}
""".strip()

                if judge_model_type == "cloud":
                    judge_url = f"{os.getenv('BENCHMARK_JUDGE_BASE_URL', 'https://api.deepseek.com/v1')}/chat/completions"
                    judge_headers = {
                        "Authorization": f"Bearer {judge_api_key}",
                        "Content-Type": "application/json",
                    }
                    judge_model = BENCHMARK_JUDGE_MODEL
                else:
                    # Parse local ollama url from query_llm_binding_host
                    judge_url = f"{os.getenv('BENCHMARK_LOCAL_JUDGE_URL', 'http://127.0.0.1:11434/v1')}/chat/completions"
                    judge_headers = {"Content-Type": "application/json"}
                    judge_model = os.getenv("BENCHMARK_LOCAL_JUDGE_MODEL", benchmark_model_config.model)

                response = requests.post(
                    judge_url,
                    headers=judge_headers,
                    json={
                        "model": judge_model,
                        "messages": [
                            {
                                "role": "system",
                                "content": "你是一位严格的 Benchmark 评测裁判，请用中文思维理解和打分。",
                            },
                            {"role": "user", "content": prompt},
                        ],
                        "response_format": {"type": "json_object"},
                    },
                    timeout=float(benchmark_judge_timeout),
                )
                response.raise_for_status()
                payload = response.json()
                content = (
                    payload.get("choices", [{}])[0]
                    .get("message", {})
                    .get("content", "{}")
                ) or "{}"
                payload = json.loads(content)
                judged: dict[str, int] = {}
                for mode in mode_answers:
                    raw_score = payload.get(mode)
                    if raw_score not in SCORE_LABELS:
                        raise RuntimeError(
                            f"Judge returned invalid score for {mode}: {raw_score!r}"
                        )
                    judged[mode] = int(raw_score)
                return judged

            async def collect_benchmark_stream_metrics(response: Any) -> tuple[str, float, float]:
                import time as time_module

                started_at = time_module.perf_counter()

                if isinstance(response, str):
                    elapsed = time_module.perf_counter() - started_at
                    return response, elapsed, elapsed

                first_chunk_time: float | None = None
                chunks: list[str] = []

                try:
                    while True:
                        try:
                            candidate = await response.__anext__()
                        except StopAsyncIteration:
                            break
                        if candidate:
                            first_chunk_time = time_module.perf_counter()
                            chunks.append(str(candidate))
                            break

                    async for chunk in response:
                        if chunk:
                            chunks.append(str(chunk))
                except Exception:
                    raise

                full_elapsed = time_module.perf_counter() - started_at
                ttft_elapsed = (
                    first_chunk_time - started_at
                    if first_chunk_time is not None
                    else full_elapsed
                )
                return "".join(chunks), ttft_elapsed, full_elapsed

            async def executor_async(row, rag_instance):
                question = row.get("question", "")
                if not str(question).strip():
                    raise ValueError(
                        "Benchmark source row is missing a question after header normalization."
                    )

                mode_answers: dict[str, str] = {}
                mode_metrics: dict[str, dict[str, float]] = {}
                mode_scores: dict[str, int] = {}
                judge_error_message: str | None = None
                expected_source_ids = parse_evidence_source_ids(row.get("evidence"))

                async def fetch_mode_context(mode_name: str) -> str:
                    context_response = await asyncio.wait_for(
                        rag_instance.aquery(
                            question,
                            param=QueryParam(
                                mode=getattr(SearchMode, mode_name),
                                stream=False,
                                only_need_context=True,
                                top_k=shared_query_profile["top_k"],
                                response_type=shared_query_profile["response_type"],
                                max_token_for_text_unit=shared_query_profile["max_token_for_text_unit"],
                                max_token_for_node_context=shared_query_profile["max_token_for_node_context"],
                            ),
                        ),
                        timeout=benchmark_call_timeout,
                    )
                    return str(context_response or "")

                async def run_single_mode(mode_name: str):
                    if benchmark_state.should_stop():
                        return mode_name, {"stop_requested": True}
                    recall_rate = 0.0
                    try:
                        try:
                            context_text = await fetch_mode_context(mode_name)
                            retrieved_source_ids = parse_context_source_ids(context_text)
                            recall_rate = compute_recall_rate(
                                expected_source_ids,
                                retrieved_source_ids,
                            )
                        except Exception as context_err:
                            logging.warning(
                                "Benchmark context recall collection failed for %s: %s",
                                mode_name,
                                context_err,
                            )
                        response = await asyncio.wait_for(
                            rag_instance.aquery(
                                question,
                                param=QueryParam(
                                    mode=getattr(SearchMode, mode_name),
                                    stream=True,
                                    top_k=shared_query_profile["top_k"],
                                    response_type=shared_query_profile["response_type"],
                                    max_token_for_text_unit=shared_query_profile["max_token_for_text_unit"],
                                    max_token_for_node_context=shared_query_profile["max_token_for_node_context"],
                                ),
                            ),
                            timeout=benchmark_call_timeout,
                        )
                        model_answer, ttft_elapsed, full_elapsed = (
                            await collect_benchmark_stream_metrics(response)
                        )
                        return mode_name, {
                            "answer": model_answer,
                            "metrics": {
                                "response_time_ms": round(ttft_elapsed * 1000),
                                "ttft_total_s": ttft_elapsed,
                                "full_s": full_elapsed,
                                "recall_rate": recall_rate,
                            },
                        }
                    except Exception as e:
                        return mode_name, {
                            "answer": f"[ERROR] {format_benchmark_exception(e)}",
                            "metrics": {"recall_rate": 0.0},
                            "score": -1,
                        }

                mode_tasks = [run_single_mode(m) for m in BENCHMARK_QUERY_MODES]
                mode_results = await asyncio.gather(*mode_tasks)

                for mode_name, mode_res in mode_results:
                    if mode_res.get("stop_requested"):
                        return {"stop_requested": True}
                    mode_answers[mode_name] = mode_res["answer"]
                    mode_metrics[mode_name] = mode_res["metrics"]
                    if "score" in mode_res:
                        mode_scores[mode_name] = mode_res["score"]

                if benchmark_state.should_stop():
                    return {"stop_requested": True}

                judge_ready_answers = {
                    mode_name: answer
                    for mode_name, answer in mode_answers.items()
                    if not answer.startswith("[ERROR]")
                }

                if benchmark_state.should_stop():
                    return {"stop_requested": True}

                if judge_ready_answers:
                    try:
                        mode_scores.update(
                            await asyncio.wait_for(
                                asyncio.to_thread(
                                    judge_answers, 
                                    row, 
                                    judge_ready_answers,
                                    request.judge_model_type
                                ),
                                timeout=benchmark_judge_timeout,
                            )
                        )
                    except TimeoutError:
                        if benchmark_state.should_stop():
                            return {"stop_requested": True}
                        judge_error_message = (
                            f"Benchmark judge timed out after {benchmark_judge_timeout}s"
                        )
                    except Exception as e:
                        if benchmark_state.should_stop():
                            return {"stop_requested": True}
                        judge_error_message = (
                            "Benchmark judge failed: "
                            f"{format_benchmark_exception(e)}"
                        )

                for mode_name in BENCHMARK_QUERY_MODES:
                    mode_scores.setdefault(mode_name, -1)

                primary_mode = "graph_text_hybrid"
                primary_answer = mode_answers.get(primary_mode, "")
                primary_metrics = mode_metrics.get(primary_mode, {})

                return {
                    "faultcase_fast_stream_plain": primary_answer,
                    "raw_score": mode_scores.get(primary_mode, -1),
                    "metrics": primary_metrics,
                    "mode_answers": mode_answers,
                    "mode_scores": mode_scores,
                    "mode_metrics": mode_metrics,
                    "primary_mode": primary_mode,
                    "error_message": judge_error_message,
                }

            def run_with_score_handling():
                benchmark_loop = asyncio.new_event_loop()
                asyncio.set_event_loop(benchmark_loop)
                try:
                    rag_instance = build_benchmark_rag_instance()

                    # Use concurrent arun_current_flow for questions level concurrency
                    results = benchmark_loop.run_until_complete(
                        ship_benchmark_adapter.arun_current_flow(
                            run_id,
                            lambda row: executor_async(row, rag_instance),
                            rows,
                            on_result=result_writer.write_item,
                            concurrency=5, # default concurrency
                        )
                    )
                    return results
                except ValueError as e:
                    if benchmark_state.state == BenchmarkState.stopping:
                        benchmark_state.finalize_stop_if_worker_finished()
                    else:
                        benchmark_state.fail(str(e))
                    return []
                except Exception as e:
                    if benchmark_state.state == BenchmarkState.stopping:
                        benchmark_state.finalize_stop_if_worker_finished()
                    else:
                        benchmark_state.fail(str(e))
                    return []
                finally:
                    result_writer.close()
                    benchmark_loop.close()
                    benchmark_state.worker_finished(threading.current_thread())

            thread = threading.Thread(target=run_with_score_handling, daemon=True)
            benchmark_state.register_worker(thread)
            thread.start()

            initial_state = benchmark_state.state
            if initial_state == BenchmarkState.idle:
                initial_state = BenchmarkState.starting

            return BenchmarkRunResponse(
                status="started",
                run_id=run_id,
                state=initial_state,
                selected_model=selected_model_id,
                message=f"Benchmark run {run_id} started with {len(rows)} questions using {selected_model_id}{runtime_note}.",
            )

        except ValueError as e:
            raise HTTPException(status_code=409, detail=str(e))
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.get(
        "/benchmark/status",
        response_model=BenchmarkPageSnapshot,
        dependencies=[Depends(optional_api_key)],
    )
    async def get_benchmark_status():
        """
        Get the full normalized benchmark snapshot.

        Returns the canonical page snapshot directly from the state container,
        including terminal snapshot for completed/stopped/failed states.
        """
        benchmark_state.finalize_stop_if_worker_finished()
        return benchmark_state.to_page_snapshot()

    @app.post(
        "/benchmark/stop",
        response_model=BenchmarkStopResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def stop_benchmark_run(request: BenchmarkStopRequest | None = None):
        """
        Stop the current benchmark run.

        Transitions active runs (starting/running) into stopping state,
        then finalizes to stopped state.
        """
        try:
            if not benchmark_state.can_stop():
                raise HTTPException(
                    status_code=409,
                    detail=f"Cannot stop benchmark: current state is {benchmark_state.state.value}",
                )

            benchmark_state.stop(request.reason if request is not None else None)
            finalized = benchmark_state.finalize_stop_if_worker_finished()

            return BenchmarkStopResponse(
                status="stopped" if finalized else "stopping",
                state=benchmark_state.state,
                message=(
                    (request.reason if request is not None else None)
                    or "Benchmark run stopped by user"
                    if finalized
                    else (request.reason if request is not None else None)
                    or "Stop requested; current active question will finish before finalizing."
                ),
            )

        except ValueError as e:
            raise HTTPException(status_code=409, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    @app.post(
        "/benchmark/reset",
        response_model=BenchmarkResetResponse,
        dependencies=[Depends(optional_api_key)],
    )
    async def reset_benchmark():
        """
        Reset benchmark to canonical idle state.

        Only allowed from terminal states (completed/stopped/failed).
        Clears all terminal snapshot data.
        """
        try:
            if not benchmark_state.can_reset():
                raise HTTPException(
                    status_code=409,
                    detail=f"Cannot reset benchmark: current state is {benchmark_state.state.value}. Reset only allowed from terminal states.",
                )

            benchmark_state.reset()

            return BenchmarkResetResponse(
                status="reset",
                state=benchmark_state.state,
                message="Benchmark reset to idle state",
            )

        except ValueError as e:
            raise HTTPException(status_code=409, detail=str(e))
        except Exception as e:
            raise HTTPException(status_code=500, detail=str(e))

    # ============================================================
    # End of Benchmark Endpoints
    # ============================================================

    # Raw Chunk Registry Initialization (merged with benchmark endpoints)
    from minirag.api.raw_chunk_registry import RawChunkRegistry
    from minirag.api.raw_chunk_endpoints import register_raw_chunk_endpoints

    # Initialize raw chunk registry
    raw_chunks_root = resolved_datasource.staging_root / "chunks"
    raw_chunk_registry = RawChunkRegistry(args.working_dir, str(raw_chunks_root))
    raw_chunk_registry.sync_from_disk()

    # Register raw chunk endpoints - also exported for testing
    app = register_raw_chunk_endpoints(
        app,
        raw_chunk_registry,
        str(raw_chunks_root),
        dependencies=[Depends(optional_api_key)],
    )

    # webui mount /webui/index.html
    # app.mount(
    #     "/webui",
    #     StaticFiles(
    #         directory=Path(__file__).resolve().parent / "webui" / "static", html=True
    #     ),
    #     name="webui_static",
    # )

    if args.serve_static_ui:
        # Serve the static files
        static_dir = Path(__file__).parent / "static"
        static_dir.mkdir(exist_ok=True)
        app.mount("/", StaticFiles(directory=static_dir, html=True), name="static")

    return app


def main():
    args = parse_args()
    import uvicorn

    app = create_app(args)
    display_splash_screen(args)
    uvicorn_config = {
        "app": app,
        "host": args.host,
        "port": args.port,
    }
    if args.ssl:
        uvicorn_config.update(
            {
                "ssl_certfile": args.ssl_certfile,
                "ssl_keyfile": args.ssl_keyfile,
            }
        )
    uvicorn.run(**uvicorn_config)


if __name__ == "__main__":
    main()
